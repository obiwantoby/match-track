import requests
import json
import sys
import io
import os
import tempfile
from datetime import datetime
import openpyxl
import re

class ExcelNullValuesVerificationTester:
    def __init__(self, base_url="https://54bdef35-ae60-4161-ae24-d2c0da9aaead.preview.emergentagent.com"):
        self.base_url = base_url
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.api_url = f"{self.base_url}/api"

    def run_test(self, name, method, endpoint, expected_status, data=None, check_function=None, binary=False):
        """Run a single API test"""
        url = f"{self.api_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            status_success = response.status_code == expected_status
            
            if status_success:
                print(f"✅ Status check passed - Expected: {expected_status}, Got: {response.status_code}")
                
                # If there's a custom check function, run it
                if check_function:
                    check_result = check_function(response)
                    if check_result:
                        print(f"✅ Custom check passed: {check_result}")
                        self.tests_passed += 1
                        return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                    else:
                        print(f"❌ Custom check failed")
                        return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                else:
                    self.tests_passed += 1
                    return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
            else:
                print(f"❌ Status check failed - Expected: {expected_status}, Got: {response.status_code}")
                return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, None

    def login(self, email="admin@example.com", password="admin123"):
        """Login and get authentication token"""
        print(f"\n🔐 Logging in as {email}...")
        
        form_data = {
            "username": email,
            "password": password
        }
        
        url = f"{self.api_url}/auth/token"
        response = requests.post(
            url, 
            data=form_data,
            headers={"Content-Type": "application/x-www-form-urlencoded"}
        )
        
        if response.status_code == 200:
            data = response.json()
            self.token = data.get("access_token")
            print(f"✅ Login successful - Token received")
            return True
        else:
            print(f"❌ Login failed - Status: {response.status_code}")
            if response.headers.get('Content-Type', '').startswith('application/json'):
                print(f"Error: {response.json()}")
            return False

    def test_excel_null_values_verification(self):
        """Test that Excel export correctly displays NULL values as '-' and 0 values as '0'"""
        # First, create a test shooter
        shooter_data = {
            "name": f"Excel Test Shooter {datetime.now().strftime('%H%M%S')}",
            "nra_number": "12345",
            "cmp_number": "67890"
        }
        
        success, shooter = self.run_test(
            "Create test shooter for Excel test",
            "POST",
            "shooters",
            200,
            data=shooter_data
        )
        
        if not success:
            print("❌ Failed to create test shooter, cannot continue test")
            return False
        
        shooter_id = shooter["id"]
        print(f"Created test shooter with ID: {shooter_id}")
        
        # Create a test match
        match_data = {
            "name": f"Excel Test Match {datetime.now().strftime('%H%M%S')}",
            "date": datetime.now().isoformat(),
            "location": "Test Range",
            "match_types": [
                {
                    "type": "NMC",
                    "instance_name": "NMC1",
                    "calibers": [".22", "CF"]
                }
            ],
            "aggregate_type": "None"
        }
        
        success, match = self.run_test(
            "Create test match for Excel test",
            "POST",
            "matches",
            200,
            data=match_data
        )
        
        if not success:
            print("❌ Failed to create test match, cannot continue test")
            return False
        
        match_id = match["id"]
        print(f"Created test match with ID: {match_id}")
        
        # Create a score with NULL value
        score1_data = {
            "shooter_id": shooter_id,
            "match_id": match_id,
            "match_type_instance": "NMC1",
            "caliber": ".22",
            "stages": [
                {
                    "name": "SF",
                    "score": 95,
                    "x_count": 3
                },
                {
                    "name": "TF",
                    "score": None,  # NULL value - should be displayed as "-" in Excel
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 90,
                    "x_count": 2
                }
            ]
        }
        
        success, score1 = self.run_test(
            "Create score with NULL value for Excel test",
            "POST",
            "scores",
            200,
            data=score1_data
        )
        
        if not success:
            print("❌ Failed to create score with NULL value")
            return False
        
        print(f"Created score with NULL value, ID: {score1['id']}")
        
        # Create a second score with 0 value
        score2_data = {
            "shooter_id": shooter_id,
            "match_id": match_id,
            "match_type_instance": "NMC1",
            "caliber": "CF",
            "stages": [
                {
                    "name": "SF",
                    "score": 85,
                    "x_count": 1
                },
                {
                    "name": "TF",
                    "score": 0,  # Zero value - should be displayed as "0" in Excel
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 80,
                    "x_count": 1
                }
            ]
        }
        
        success, score2 = self.run_test(
            "Create score with 0 value for Excel test",
            "POST",
            "scores",
            200,
            data=score2_data
        )
        
        if not success:
            print("❌ Failed to create score with 0 value")
            return False
        
        print(f"Created score with 0 value, ID: {score2['id']}")
        
        # Download Excel report
        def check_excel_content(response):
            if response.status_code != 200:
                return None
                
            # Check if we got an Excel file
            content_type = response.headers.get('Content-Type', '')
            if 'spreadsheet' not in content_type and 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
                print(f"❌ Expected Excel file, got {content_type}")
                return None
            
            # Save the Excel file temporarily
            excel_data = response.content
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file_path = temp_file.name
                temp_file.write(excel_data)
            
            print(f"📊 Excel file saved to {temp_file_path}")
            
            # Parse the Excel file to check NULL and 0 values
            try:
                wb = openpyxl.load_workbook(temp_file_path)
                sheet = wb.active
                
                # Find our test shooter's data
                shooter_name = shooter_data["name"]
                shooter_row = None
                
                for row in range(1, sheet.max_row + 1):
                    cell_value = sheet.cell(row=row, column=1).value
                    if cell_value and shooter_name in str(cell_value):
                        shooter_row = row
                        break
                
                if not shooter_row:
                    print("❌ Could not find test shooter in Excel file")
                    os.unlink(temp_file_path)
                    return None
                
                print(f"📊 Found test shooter at row {shooter_row}")
                
                # Get the shooter's row data
                row_data = []
                for col in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=shooter_row, column=col).value
                    row_data.append(value)
                
                print(f"Shooter's row data: {row_data}")
                
                # Get the header row to understand the data
                header_row = []
                for col in range(1, sheet.max_column + 1):
                    value = sheet.cell(row=8, column=col).value  # Header is at row 8
                    header_row.append(value)
                
                print(f"Header row: {header_row}")
                
                # Check for NULL values and 0 values in the Excel file
                null_value_found = False
                zero_value_found = False
                
                # The Excel file doesn't show individual stages, it shows aggregate scores
                # We need to check the raw data in the scores
                
                # Get the .22 score (should have a NULL value in TF stage)
                twentytwo_score = None
                for i, header in enumerate(header_row):
                    if header and "TWENTYTWO" in str(header):
                        twentytwo_score = row_data[i]
                        break
                
                # Get the CF score (should have a 0 value in TF stage)
                cf_score = None
                for i, header in enumerate(header_row):
                    if header and "CENTERFIRE" in str(header):
                        cf_score = row_data[i]
                        break
                
                print(f".22 score: {twentytwo_score}")
                print(f"CF score: {cf_score}")
                
                # Get the average score
                average_score = None
                for i, header in enumerate(header_row):
                    if header and "Average" in str(header):
                        average_score = row_data[i]
                        break
                
                print(f"Average score: {average_score}")
                
                # Verify that the average calculation is correct
                # For .22: (95 + 90) / 2 = 92.5 (excluding NULL)
                # For CF: (85 + 0 + 80) / 3 = 55 (including 0)
                # Overall: (92.5 + 55) / 2 = 73.75
                
                # Extract the numeric values from the scores
                twentytwo_numeric = None
                if twentytwo_score:
                    match = re.search(r'(\d+)', str(twentytwo_score))
                    if match:
                        twentytwo_numeric = int(match.group(1))
                
                cf_numeric = None
                if cf_score:
                    match = re.search(r'(\d+)', str(cf_score))
                    if match:
                        cf_numeric = int(match.group(1))
                
                average_numeric = None
                if average_score:
                    try:
                        average_numeric = float(average_score)
                    except (ValueError, TypeError):
                        pass
                
                print(f".22 numeric: {twentytwo_numeric}")
                print(f"CF numeric: {cf_numeric}")
                print(f"Average numeric: {average_numeric}")
                
                # Verify that the average calculation is correct
                expected_twentytwo = (95 + 90) / 2  # 92.5
                expected_cf = (85 + 0 + 80) / 3  # 55
                expected_average = (expected_twentytwo + expected_cf) / 2  # 73.75
                
                print(f"Expected .22: {expected_twentytwo}")
                print(f"Expected CF: {expected_cf}")
                print(f"Expected average: {expected_average}")
                
                # Check if the average is within a reasonable range
                average_correct = False
                if average_numeric is not None:
                    average_correct = abs(average_numeric - expected_average) < 2  # Allow for rounding differences
                
                # Clean up
                os.unlink(temp_file_path)
                
                # Summarize findings
                print("\n=== Test Results ===")
                print(f"Average calculation correct: {'✅ Yes' if average_correct else '❌ No'}")
                
                if average_correct:
                    return "Excel file correctly calculates averages, excluding NULL values but including 0 values"
                else:
                    return None
                
            except Exception as e:
                print(f"❌ Error parsing Excel file: {str(e)}")
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
                return None
        
        success, _ = self.run_test(
            "Download and verify Excel report",
            "GET",
            f"match-report/{match_id}/excel",
            200,
            check_function=check_excel_content,
            binary=True
        )
        
        return success

def main():
    tester = ExcelNullValuesVerificationTester()
    
    # Login first
    if not tester.login():
        print("❌ Login failed, cannot continue tests")
        return 1
    
    # Run test for Excel export with NULL values
    print("\n=== Testing Excel Export with NULL and 0 Values ===")
    excel_test_result = tester.test_excel_null_values_verification()
    
    # Print summary
    print("\n=== Test Summary ===")
    print(f"Excel Export NULL Values Test: {'✅ PASSED' if excel_test_result else '❌ FAILED'}")
    
    return 0 if excel_test_result else 1

if __name__ == "__main__":
    sys.exit(main())

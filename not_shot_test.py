import requests
import json
import os
import io
import uuid
from datetime import datetime, timedelta
import openpyxl
from openpyxl.utils import get_column_letter

# Get the backend URL from the frontend .env file
def get_backend_url():
    with open('/app/frontend/.env', 'r') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                return line.strip().split('=')[1].strip('"\'')
    return None

BACKEND_URL = get_backend_url()
if not BACKEND_URL:
    raise ValueError("Could not find REACT_APP_BACKEND_URL in frontend/.env")

API_URL = f"{BACKEND_URL}/api"
print(f"Using API URL: {API_URL}")

# Test user credentials
TEST_USER = {
    "email": "admin@example.com",
    "password": "admin123"
}

def login():
    """Login and get authentication token"""
    print(f"\n🔐 Logging in as {TEST_USER['email']}...")
    
    form_data = {
        "username": TEST_USER["email"],
        "password": TEST_USER["password"]
    }
    
    url = f"{API_URL}/auth/token"
    response = requests.post(
        url, 
        data=form_data,
        headers={"Content-Type": "application/x-www-form-urlencoded"}
    )
    
    if response.status_code == 200:
        data = response.json()
        token = data.get("access_token")
        print(f"✅ Login successful - Token received")
        return token
    else:
        print(f"❌ Login failed - Status: {response.status_code}")
        if response.headers.get('Content-Type', '').startswith('application/json'):
            print(f"Error: {response.json()}")
        return None

def create_test_shooter(token):
    """Create a test shooter for the Excel export test"""
    shooter_data = {
        "name": f"Not Shot Test Shooter {datetime.now().strftime('%H%M%S')}",
        "nra_number": "12345",
        "cmp_number": "67890"
    }
    
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.post(f"{API_URL}/shooters", json=shooter_data, headers=headers)
    
    if response.status_code == 200:
        shooter = response.json()
        shooter_id = shooter["id"]
        print(f"✅ Created test shooter with ID: {shooter_id}")
        return shooter_id, shooter_data["name"]
    else:
        print(f"❌ Failed to create test shooter: {response.status_code} - {response.text}")
        return None, None

def create_test_match(token):
    """Create a test match with multiple match types and calibers"""
    match_data = {
        "name": f"Not Shot Test Match {datetime.now().strftime('%H%M%S')}",
        "date": datetime.now().isoformat(),
        "location": "Test Range",
        "match_types": [
            {
                "type": "NMC",
                "instance_name": "NMC1",
                "calibers": [".22", "CF", ".45"]
            },
            {
                "type": "600",
                "instance_name": "600_1",
                "calibers": [".22", "CF"]
            }
        ],
        "aggregate_type": "None"
    }
    
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.post(f"{API_URL}/matches", json=match_data, headers=headers)
    
    if response.status_code == 200:
        match = response.json()
        match_id = match["id"]
        print(f"✅ Created test match with ID: {match_id}")
        return match_id
    else:
        print(f"❌ Failed to create test match: {response.status_code} - {response.text}")
        return None

def create_not_shot_scores(token, shooter_id, match_id):
    """Create scores with not_shot flag set to true"""
    headers = {"Authorization": f"Bearer {token}"}
    
    # Create score for NMC1 - .22 (regular score)
    nmc_22_score = {
        "shooter_id": shooter_id,
        "match_id": match_id,
        "caliber": ".22",
        "match_type_instance": "NMC1",
        "stages": [
            {"name": "SF", "score": 95, "x_count": 3},
            {"name": "TF", "score": 92, "x_count": 2},
            {"name": "RF", "score": 90, "x_count": 1}
        ]
    }
    
    # Create score for NMC1 - CF (not shot)
    nmc_cf_score = {
        "shooter_id": shooter_id,
        "match_id": match_id,
        "caliber": "CF",
        "match_type_instance": "NMC1",
        "stages": [
            {"name": "SF", "score": None, "x_count": None},
            {"name": "TF", "score": None, "x_count": None},
            {"name": "RF", "score": None, "x_count": None}
        ],
        "not_shot": True
    }
    
    # Create score for NMC1 - .45 (not shot)
    nmc_45_score = {
        "shooter_id": shooter_id,
        "match_id": match_id,
        "caliber": ".45",
        "match_type_instance": "NMC1",
        "stages": [
            {"name": "SF", "score": None, "x_count": None},
            {"name": "TF", "score": None, "x_count": None},
            {"name": "RF", "score": None, "x_count": None}
        ],
        "not_shot": True
    }
    
    # Create score for 600_1 - .22 (regular score with 0)
    six_22_score = {
        "shooter_id": shooter_id,
        "match_id": match_id,
        "caliber": ".22",
        "match_type_instance": "600_1",
        "stages": [
            {"name": "SF1", "score": 98, "x_count": 4},
            {"name": "SF2", "score": 97, "x_count": 3},
            {"name": "TF1", "score": 0, "x_count": 0},  # Zero score
            {"name": "TF2", "score": 95, "x_count": 2},
            {"name": "RF1", "score": 94, "x_count": 1},
            {"name": "RF2", "score": 93, "x_count": 0}
        ]
    }
    
    # Create score for 600_1 - CF (not shot)
    six_cf_score = {
        "shooter_id": shooter_id,
        "match_id": match_id,
        "caliber": "CF",
        "match_type_instance": "600_1",
        "stages": [
            {"name": "SF1", "score": None, "x_count": None},
            {"name": "SF2", "score": None, "x_count": None},
            {"name": "TF1", "score": None, "x_count": None},
            {"name": "TF2", "score": None, "x_count": None},
            {"name": "RF1", "score": None, "x_count": None},
            {"name": "RF2", "score": None, "x_count": None}
        ],
        "not_shot": True
    }
    
    scores = []
    for score_data in [nmc_22_score, nmc_cf_score, nmc_45_score, six_22_score, six_cf_score]:
        response = requests.post(f"{API_URL}/scores", json=score_data, headers=headers)
        if response.status_code == 200:
            score = response.json()
            caliber = score_data["caliber"]
            match_type = score_data["match_type_instance"]
            not_shot = score_data.get("not_shot", False)
            status = "not shot" if not_shot else "regular score"
            print(f"✅ Created {match_type} - {caliber} score ({status})")
            scores.append(score)
        else:
            print(f"❌ Failed to create score: {response.status_code} - {response.text}")
    
    return scores

def verify_excel_export(token, match_id, shooter_name):
    """Download and verify the Excel export for not_shot indicators"""
    headers = {"Authorization": f"Bearer {token}"}
    response = requests.get(f"{API_URL}/match-report/{match_id}/excel", headers=headers)
    
    if response.status_code != 200:
        print(f"❌ Failed to download Excel report: {response.status_code} - {response.text}")
        return False
    
    # Save the Excel file temporarily
    excel_data = response.content
    temp_file = f"/tmp/not_shot_test_{datetime.now().strftime('%H%M%S')}.xlsx"
    
    with open(temp_file, 'wb') as f:
        f.write(excel_data)
    
    print(f"📊 Excel file saved to {temp_file}")
    
    # Parse the Excel file to check not_shot indicators
    try:
        wb = openpyxl.load_workbook(temp_file)
        
        # Check the summary sheet first
        summary_sheet = wb.active
        print(f"\n🔍 Checking summary sheet: {summary_sheet.title}")
        
        # Find our test shooter's data
        shooter_row = None
        for row in range(1, summary_sheet.max_row + 1):
            cell_value = summary_sheet.cell(row=row, column=1).value
            if cell_value and shooter_name in str(cell_value):
                shooter_row = row
                break
        
        if not shooter_row:
            print("❌ Could not find test shooter in Excel summary sheet")
            return False
        
        print(f"✅ Found test shooter at row {shooter_row} in summary sheet")
        
        # Check for "-" in not shot match columns
        not_shot_dash_found = False
        for col in range(3, summary_sheet.max_column + 1):
            header = summary_sheet.cell(row=8, column=col).value
            cell_value = summary_sheet.cell(row=shooter_row, column=col).value
            
            if header and ("NMC1 (CF)" in str(header) or "NMC1 (.45)" in str(header) or "600_1 (CF)" in str(header)):
                if cell_value == "-":
                    not_shot_dash_found = True
                    print(f"✅ Found not_shot displayed as '-' in summary sheet for {header}")
        
        if not not_shot_dash_found:
            print("❌ Not shot matches are not displayed as '-' in summary sheet")
        
        # Now check the shooter's detail sheet
        detail_sheet_found = False
        not_shot_indicators_found = False
        not_shot_total_dash_found = False
        
        for sheet_name in wb.sheetnames:
            if shooter_name[:28] in sheet_name:  # Sheet names are limited to 31 chars
                detail_sheet = wb[sheet_name]
                detail_sheet_found = True
                print(f"\n🔍 Checking detail sheet: {sheet_name}")
                
                # Find all section headers (match type - caliber)
                section_headers = []
                for row in range(1, detail_sheet.max_row + 1):
                    cell = detail_sheet.cell(row=row, column=1)
                    # Check for section headers (they have a gray background)
                    if cell.fill.start_color.rgb == "FFD9D9D9" and cell.font.bold:
                        section_headers.append((row, cell.value))
                
                print(f"Found {len(section_headers)} section headers")
                
                # Check for "Not Shot" indicators
                not_shot_indicators = []
                for row in range(1, detail_sheet.max_row + 1):
                    cell = detail_sheet.cell(row=row, column=1)
                    if cell.value == "Not Shot":
                        not_shot_indicators.append(row)
                        not_shot_indicators_found = True
                        
                        # Check if it's red and merged
                        if cell.font.color.rgb == "FFFF0000":
                            print(f"✅ 'Not Shot' indicator at row {row} is correctly displayed in red")
                        else:
                            print(f"❌ 'Not Shot' indicator at row {row} is not red")
                        
                        # Check if it's merged across columns A-C
                        merged_ranges = [r for r in detail_sheet.merged_cells.ranges if row in range(r.min_row, r.max_row + 1)]
                        if merged_ranges and any(r.min_col == 1 and r.max_col == 3 for r in merged_ranges):
                            print(f"✅ 'Not Shot' indicator at row {row} is correctly merged across columns A-C")
                        else:
                            print(f"❌ 'Not Shot' indicator at row {row} is not properly merged across columns A-C")
                        
                        # Check if the total row below has dashes
                        total_row = row + 1
                        if detail_sheet.cell(row=total_row, column=1).value == "Total":
                            if detail_sheet.cell(row=total_row, column=2).value == "-" and detail_sheet.cell(row=total_row, column=3).value == "-":
                                not_shot_total_dash_found = True
                                print(f"✅ Total row for not_shot match at row {total_row} correctly displays '-' for score and X count")
                            else:
                                print(f"❌ Total row for not_shot match at row {total_row} does not display '-' for score and X count")
                
                print(f"Found {len(not_shot_indicators)} 'Not Shot' indicators")
                
                break
        
        # Clean up
        os.remove(temp_file)
        
        # Summarize findings
        if detail_sheet_found and not_shot_indicators_found and not_shot_dash_found and not_shot_total_dash_found:
            print("\n✅ Excel export correctly handles not_shot matches:")
            print("1. Not shot matches are displayed as '-' in the summary sheet")
            print("2. Not shot matches have 'Not Shot' indicator in red in detail sheets")
            print("3. Total rows for not shot matches display '-' for score and X count")
            return True
        else:
            print("\n❌ Issues found with not_shot handling in Excel export:")
            if not detail_sheet_found:
                print("- Shooter detail sheet not found")
            if not not_shot_indicators_found:
                print("- 'Not Shot' indicators not found in detail sheets")
            if not not_shot_dash_found:
                print("- Not shot matches not displayed as '-' in summary sheet")
            if not not_shot_total_dash_found:
                print("- Total rows for not shot matches don't display '-' for score and X count")
            return False
    
    except Exception as e:
        print(f"❌ Error parsing Excel file: {str(e)}")
        return False

def main():
    # Login
    token = login()
    if not token:
        print("❌ Login failed, cannot continue test")
        return 1
    
    # Create test shooter
    shooter_id, shooter_name = create_test_shooter(token)
    if not shooter_id:
        print("❌ Failed to create test shooter, cannot continue test")
        return 1
    
    # Create test match
    match_id = create_test_match(token)
    if not match_id:
        print("❌ Failed to create test match, cannot continue test")
        return 1
    
    # Create scores with not_shot flag
    scores = create_not_shot_scores(token, shooter_id, match_id)
    if not scores:
        print("❌ Failed to create scores, cannot continue test")
        return 1
    
    # Verify Excel export
    success = verify_excel_export(token, match_id, shooter_name)
    
    if success:
        print("\n✅ ALL TESTS PASSED - Excel export correctly handles not_shot matches")
    else:
        print("\n❌ SOME TESTS FAILED - Issues found with not_shot handling in Excel export")
    
    return 0 if success else 1

if __name__ == "__main__":
    main()
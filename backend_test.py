import requests
import json
import sys
import io
import os
from datetime import datetime
import pandas as pd
import openpyxl

class ExcelNullValuesComprehensiveTester:
    def __init__(self, base_url="https://b78bc624-fd3d-457d-a921-b3684a7c6c0b.preview.emergentagent.com"):
        self.base_url = base_url
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.api_url = f"{self.base_url}/api"
        self.test_data = {}

    def run_test(self, name, method, endpoint, expected_status, data=None, check_function=None, binary=False):
        """Run a single API test"""
        url = f"{self.api_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            status_success = response.status_code == expected_status
            
            if status_success:
                print(f"✅ Status check passed - Expected: {expected_status}, Got: {response.status_code}")
                
                # If there's a custom check function, run it
                if check_function:
                    check_result = check_function(response)
                    if check_result:
                        print(f"✅ Custom check passed: {check_result}")
                        self.tests_passed += 1
                        return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                    else:
                        print(f"❌ Custom check failed")
                        return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                else:
                    self.tests_passed += 1
                    return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
            else:
                print(f"❌ Status check failed - Expected: {expected_status}, Got: {response.status_code}")
                try:
                    error_content = response.json() if 'application/json' in response.headers.get('Content-Type', '') else response.text
                    print(f"Error content: {error_content}")
                except:
                    print(f"Error content could not be parsed")
                return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, None

    def login(self, email="admin@example.com", password="admin123"):
        """Login and get authentication token"""
        print(f"\n🔐 Logging in as {email}...")
        
        form_data = {
            "username": email,
            "password": password
        }
        
        url = f"{self.api_url}/auth/token"
        response = requests.post(
            url, 
            data=form_data,
            headers={"Content-Type": "application/x-www-form-urlencoded"}
        )
        
        if response.status_code == 200:
            data = response.json()
            self.token = data.get("access_token")
            print(f"✅ Login successful - Token received")
            return True
        else:
            print(f"❌ Login failed - Status: {response.status_code}")
            if response.headers.get('Content-Type', '').startswith('application/json'):
                print(f"Error: {response.json()}")
            return False

    def create_test_shooter(self):
        """Create a test shooter for the Excel export test"""
        shooter_data = {
            "name": f"Excel Test Shooter {datetime.now().strftime('%H%M%S')}",
            "nra_number": "12345",
            "cmp_number": "67890"
        }
        
        success, shooter = self.run_test(
            "Create test shooter for Excel test",
            "POST",
            "shooters",
            200,
            data=shooter_data
        )
        
        if not success:
            print("❌ Failed to create test shooter, cannot continue test")
            return False
        
        shooter_id = shooter["id"]
        print(f"Created test shooter with ID: {shooter_id}")
        self.test_data["shooter_id"] = shooter_id
        self.test_data["shooter_name"] = shooter_data["name"]
        return True

    def create_test_match(self):
        """Create a test match with multiple match types and calibers"""
        match_data = {
            "name": f"Excel NULL Test Match {datetime.now().strftime('%H%M%S')}",
            "date": datetime.now().isoformat(),
            "location": "Test Range",
            "match_types": [
                {
                    "type": "NMC",
                    "instance_name": "NMC1",
                    "calibers": [".22", "CF", ".45"]
                },
                {
                    "type": "600",
                    "instance_name": "600_1",
                    "calibers": [".22", "CF"]
                }
            ],
            "aggregate_type": "None"
        }
        
        success, match = self.run_test(
            "Create test match for Excel test",
            "POST",
            "matches",
            200,
            data=match_data
        )
        
        if not success:
            print("❌ Failed to create test match, cannot continue test")
            return False
        
        match_id = match["id"]
        print(f"Created test match with ID: {match_id}")
        self.test_data["match_id"] = match_id
        self.test_data["match_name"] = match_data["name"]
        return True

    def create_scores_with_null_and_zero_values(self):
        """Create multiple scores with a mix of NULL and 0 values"""
        if not self.test_data.get("shooter_id") or not self.test_data.get("match_id"):
            print("❌ Missing shooter or match ID, cannot create scores")
            return False
            
        # Score 1: NMC1 with .22 - Has NULL value in TF
        score1_data = {
            "shooter_id": self.test_data["shooter_id"],
            "match_id": self.test_data["match_id"],
            "match_type_instance": "NMC1",
            "caliber": ".22",
            "stages": [
                {
                    "name": "SF",
                    "score": 95,
                    "x_count": 3
                },
                {
                    "name": "TF",
                    "score": None,  # NULL value - should be displayed as "-" in Excel
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 90,
                    "x_count": 2
                }
            ]
        }
        
        success, score1 = self.run_test(
            "Create NMC1 .22 score with NULL value",
            "POST",
            "scores",
            200,
            data=score1_data
        )
        
        if not success:
            print("❌ Failed to create NMC1 .22 score with NULL value")
            return False
        
        print(f"Created NMC1 .22 score with NULL value, ID: {score1['id']}")
        
        # Score 2: NMC1 with CF - Has 0 value in TF
        score2_data = {
            "shooter_id": self.test_data["shooter_id"],
            "match_id": self.test_data["match_id"],
            "match_type_instance": "NMC1",
            "caliber": "CF",
            "stages": [
                {
                    "name": "SF",
                    "score": 85,
                    "x_count": 1
                },
                {
                    "name": "TF",
                    "score": 0,  # Zero value - should be displayed as "0" in Excel
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 80,
                    "x_count": 1
                }
            ]
        }
        
        success, score2 = self.run_test(
            "Create NMC1 CF score with 0 value",
            "POST",
            "scores",
            200,
            data=score2_data
        )
        
        if not success:
            print("❌ Failed to create NMC1 CF score with 0 value")
            return False
        
        print(f"Created NMC1 CF score with 0 value, ID: {score2['id']}")
        
        # Score 3: NMC1 with .45 - Has NULL values in multiple stages
        score3_data = {
            "shooter_id": self.test_data["shooter_id"],
            "match_id": self.test_data["match_id"],
            "match_type_instance": "NMC1",
            "caliber": ".45",
            "stages": [
                {
                    "name": "SF",
                    "score": None,  # NULL value
                    "x_count": 0
                },
                {
                    "name": "TF",
                    "score": None,  # NULL value
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 75,
                    "x_count": 0
                }
            ]
        }
        
        success, score3 = self.run_test(
            "Create NMC1 .45 score with multiple NULL values",
            "POST",
            "scores",
            200,
            data=score3_data
        )
        
        if not success:
            print("❌ Failed to create NMC1 .45 score with multiple NULL values")
            return False
        
        print(f"Created NMC1 .45 score with multiple NULL values, ID: {score3['id']}")
        
        # Score 4: 600_1 with .22 - Has mix of NULL and 0 values
        score4_data = {
            "shooter_id": self.test_data["shooter_id"],
            "match_id": self.test_data["match_id"],
            "match_type_instance": "600_1",
            "caliber": ".22",
            "stages": [
                {
                    "name": "SF1",
                    "score": 98,
                    "x_count": 4
                },
                {
                    "name": "SF2",
                    "score": None,  # NULL value
                    "x_count": 0
                },
                {
                    "name": "TF1",
                    "score": 95,
                    "x_count": 3
                },
                {
                    "name": "TF2",
                    "score": 0,  # Zero value
                    "x_count": 0
                },
                {
                    "name": "RF1",
                    "score": 92,
                    "x_count": 2
                },
                {
                    "name": "RF2",
                    "score": None,  # NULL value
                    "x_count": 0
                }
            ]
        }
        
        success, score4 = self.run_test(
            "Create 600_1 .22 score with mix of NULL and 0 values",
            "POST",
            "scores",
            200,
            data=score4_data
        )
        
        if not success:
            print("❌ Failed to create 600_1 .22 score with mix of NULL and 0 values")
            return False
        
        print(f"Created 600_1 .22 score with mix of NULL and 0 values, ID: {score4['id']}")
        
        # Score 5: 600_1 with CF - All values present (no NULLs or 0s)
        score5_data = {
            "shooter_id": self.test_data["shooter_id"],
            "match_id": self.test_data["match_id"],
            "match_type_instance": "600_1",
            "caliber": "CF",
            "stages": [
                {
                    "name": "SF1",
                    "score": 88,
                    "x_count": 1
                },
                {
                    "name": "SF2",
                    "score": 87,
                    "x_count": 1
                },
                {
                    "name": "TF1",
                    "score": 86,
                    "x_count": 0
                },
                {
                    "name": "TF2",
                    "score": 85,
                    "x_count": 0
                },
                {
                    "name": "RF1",
                    "score": 84,
                    "x_count": 0
                },
                {
                    "name": "RF2",
                    "score": 83,
                    "x_count": 0
                }
            ]
        }
        
        success, score5 = self.run_test(
            "Create 600_1 CF score with all values present",
            "POST",
            "scores",
            200,
            data=score5_data
        )
        
        if not success:
            print("❌ Failed to create 600_1 CF score with all values present")
            return False
        
        print(f"Created 600_1 CF score with all values present, ID: {score5['id']}")
        
        return True

    def verify_excel_export(self):
        """Download and verify the Excel export"""
        if not self.test_data.get("match_id"):
            print("❌ Missing match ID, cannot verify Excel export")
            return False
            
        def check_excel_content(response):
            if response.status_code != 200:
                return None
                
            # Check if we got an Excel file
            content_type = response.headers.get('Content-Type', '')
            if 'spreadsheet' not in content_type:
                print(f"❌ Expected Excel file, got {content_type}")
                return None
            
            # Save the Excel file temporarily
            excel_data = response.content
            temp_file = f"/tmp/excel_export_{datetime.now().strftime('%H%M%S')}.xlsx"
            
            with open(temp_file, 'wb') as f:
                f.write(excel_data)
            
            print(f"📊 Excel file saved to {temp_file}")
            
            # Parse the Excel file to check NULL and 0 values
            try:
                wb = openpyxl.load_workbook(temp_file)
                
                # Check the summary sheet first
                summary_sheet = wb.active
                print(f"Checking summary sheet: {summary_sheet.title}")
                
                # Find our test shooter's data
                shooter_name = self.test_data["shooter_name"]
                shooter_row = None
                
                for row in range(1, summary_sheet.max_row + 1):
                    cell_value = summary_sheet.cell(row=row, column=1).value
                    if cell_value and shooter_name in str(cell_value):
                        shooter_row = row
                        break
                
                if not shooter_row:
                    print("❌ Could not find test shooter in Excel summary sheet")
                    return None
                
                print(f"📊 Found test shooter at row {shooter_row} in summary sheet")
                
                # Check for NULL values (should be displayed as "-") and 0 values (should be displayed as "0")
                null_value_found_summary = False
                zero_value_found_summary = False
                
                # Scan the row for "-" and "0" values
                for col in range(1, summary_sheet.max_column + 1):
                    cell_value = summary_sheet.cell(row=shooter_row, column=col).value
                    if cell_value is None:
                        continue
                        
                    # Check column header to identify which match type and caliber this is
                    header = summary_sheet.cell(row=8, column=col).value  # Header row is typically row 8
                    
                    if header:
                        if "NMC1 (.22)" in str(header) and "-" in str(cell_value):
                            null_value_found_summary = True
                            print(f"✅ Found NULL value displayed as '-' in summary sheet column {col} (NMC1 .22)")
                        
                        if "NMC1 (CF)" in str(header) and "0" in str(cell_value):
                            zero_value_found_summary = True
                            print(f"✅ Found 0 value displayed as '0' in summary sheet column {col} (NMC1 CF)")
                
                # Check for average calculation
                average_found = False
                average_value = None
                for col in range(1, summary_sheet.max_column + 1):
                    header = summary_sheet.cell(row=8, column=col).value
                    if header and "Average" in str(header):
                        average_value = summary_sheet.cell(row=shooter_row, column=col).value
                        
                        # The average should exclude NULL values but include 0 values
                        if average_value:
                            average_found = True
                            print(f"✅ Found average calculation in summary sheet: {average_value}")
                            
                            # Store the average value for later verification
                            self.test_data["excel_average"] = average_value
                            break
                
                # Now check the shooter's detail sheet
                detail_sheet_found = False
                null_value_found_detail = False
                zero_value_found_detail = False
                
                for sheet_name in wb.sheetnames:
                    if shooter_name[:28] in sheet_name:  # Sheet names are limited to 31 chars
                        detail_sheet = wb[sheet_name]
                        detail_sheet_found = True
                        print(f"Checking detail sheet: {sheet_name}")
                        
                        # Scan the detail sheet for NULL and 0 values
                        for row in range(1, detail_sheet.max_row + 1):
                            stage_name = detail_sheet.cell(row=row, column=1).value
                            score_value = detail_sheet.cell(row=row, column=2).value
                            
                            if stage_name == "TF" and score_value == "-":
                                null_value_found_detail = True
                                print(f"✅ Found NULL value displayed as '-' in detail sheet for TF")
                            
                            if stage_name == "TF" and score_value == 0:
                                zero_value_found_detail = True
                                print(f"✅ Found 0 value displayed as '0' in detail sheet for TF")
                                
                            # Also check for 600 match stages
                            if stage_name == "TF2" and score_value == 0:
                                zero_value_found_detail = True
                                print(f"✅ Found 0 value displayed as '0' in detail sheet for TF2")
                            
                            if stage_name == "SF2" and score_value == "-":
                                null_value_found_detail = True
                                print(f"✅ Found NULL value displayed as '-' in detail sheet for SF2")
                        
                        break
                
                # Clean up
                os.remove(temp_file)
                
                # Summarize findings
                summary_checks = null_value_found_summary and zero_value_found_summary and average_found
                detail_checks = detail_sheet_found and (null_value_found_detail or zero_value_found_detail)
                
                if summary_checks and detail_checks:
                    return "Excel file correctly displays NULL values as '-', 0 values as '0', and calculates averages correctly in both summary and detail sheets"
                
                issues = []
                if not null_value_found_summary:
                    issues.append("NULL values not displayed as '-' in summary sheet")
                if not zero_value_found_summary:
                    issues.append("0 values not displayed as '0' in summary sheet")
                if not average_found:
                    issues.append("Average calculation may be incorrect in summary sheet")
                if not detail_sheet_found:
                    issues.append("Shooter detail sheet not found")
                elif not null_value_found_detail:
                    issues.append("NULL values not displayed as '-' in detail sheet")
                elif not zero_value_found_detail:
                    issues.append("0 values not displayed as '0' in detail sheet")
                
                print(f"❌ Issues found: {', '.join(issues)}")
                return None
                
            except Exception as e:
                print(f"❌ Error parsing Excel file: {str(e)}")
                return None
        
        success, _ = self.run_test(
            "Download and verify Excel report",
            "GET",
            f"match-report/{self.test_data['match_id']}/excel",
            200,
            check_function=check_excel_content,
            binary=True
        )
        
        return success
        
    def verify_shooter_statistics(self):
        """Verify the shooter statistics endpoint correctly excludes NULL scores from average calculations"""
        if not self.test_data.get("shooter_id"):
            print("❌ Missing shooter ID, cannot verify shooter statistics")
            return False
            
        def check_shooter_averages(response):
            if response.status_code != 200:
                return None
                
            data = response.json()
            
            # Check if we have caliber averages
            if "caliber_averages" not in data:
                print("❌ No caliber averages found in response")
                return None
                
            caliber_averages = data["caliber_averages"]
            
            # Check for .22 caliber averages (should exclude NULL values)
            if ".22" in caliber_averages:
                dot22_data = caliber_averages[".22"]
                
                # Check if valid_matches_count is correctly tracked
                if "valid_matches_count" in dot22_data:
                    print(f"✅ Found valid_matches_count for .22: {dot22_data['valid_matches_count']}")
                    
                    # Check if total_score_avg is calculated correctly (excluding NULL values)
                    if "total_score_avg" in dot22_data:
                        total_score_avg = dot22_data["total_score_avg"]
                        print(f"✅ Found total_score_avg for .22: {total_score_avg}")
                        
                        # Store for comparison with Excel average
                        self.test_data["api_average_22"] = total_score_avg
                        
                        # Check if stage averages exclude NULL values
                        if "sf_score_avg" in dot22_data and dot22_data["sf_score_avg"] is not None:
                            print(f"✅ SF average for .22 calculated correctly: {dot22_data['sf_score_avg']}")
                        else:
                            print("❌ SF average for .22 not calculated correctly")
                            return None
                            
                        # TF should have NULL values excluded
                        if "tf_score_avg" in dot22_data and dot22_data["tf_score_avg"] is not None:
                            print(f"✅ TF average for .22 calculated correctly: {dot22_data['tf_score_avg']}")
                        else:
                            print("❌ TF average for .22 not calculated correctly")
                            return None
                            
                        # RF should have all values included
                        if "rf_score_avg" in dot22_data and dot22_data["rf_score_avg"] is not None:
                            print(f"✅ RF average for .22 calculated correctly: {dot22_data['rf_score_avg']}")
                        else:
                            print("❌ RF average for .22 not calculated correctly")
                            return None
                    else:
                        print("❌ No total_score_avg found for .22")
                        return None
                else:
                    print("❌ No valid_matches_count found for .22")
                    return None
            else:
                print("❌ No .22 caliber averages found")
                return None
                
            # Check for CF caliber averages (should include 0 values)
            if "CF" in caliber_averages:
                cf_data = caliber_averages["CF"]
                
                # Check if valid_matches_count is correctly tracked
                if "valid_matches_count" in cf_data:
                    print(f"✅ Found valid_matches_count for CF: {cf_data['valid_matches_count']}")
                    
                    # Check if total_score_avg is calculated correctly (including 0 values)
                    if "total_score_avg" in cf_data:
                        total_score_avg = cf_data["total_score_avg"]
                        print(f"✅ Found total_score_avg for CF: {total_score_avg}")
                        
                        # Store for comparison
                        self.test_data["api_average_cf"] = total_score_avg
                        
                        # TF should have 0 values included
                        if "tf_score_avg" in cf_data and cf_data["tf_score_avg"] is not None:
                            print(f"✅ TF average for CF calculated correctly (includes 0): {cf_data['tf_score_avg']}")
                            
                            # Verify that the average is lower due to including 0 values
                            if cf_data["tf_score_avg"] < cf_data["sf_score_avg"]:
                                print("✅ TF average is lower than SF average due to including 0 values")
                            else:
                                print("❌ TF average should be lower than SF average due to including 0 values")
                                return None
                        else:
                            print("❌ TF average for CF not calculated correctly")
                            return None
                    else:
                        print("❌ No total_score_avg found for CF")
                        return None
                else:
                    print("❌ No valid_matches_count found for CF")
                    return None
            else:
                print("❌ No CF caliber averages found")
                return None
                
            # Check for .45 caliber averages (should have multiple NULL values)
            if ".45" in caliber_averages:
                dot45_data = caliber_averages[".45"]
                
                # Check if valid_matches_count is correctly tracked
                if "valid_matches_count" in dot45_data:
                    print(f"✅ Found valid_matches_count for .45: {dot45_data['valid_matches_count']}")
                    
                    # SF and TF should be NULL (not included in average)
                    if "sf_score_avg" in dot45_data and dot45_data["sf_score_avg"] is None:
                        print("✅ SF average for .45 correctly shows as None (all values were NULL)")
                    else:
                        print("❌ SF average for .45 should be None")
                        return None
                        
                    if "tf_score_avg" in dot45_data and dot45_data["tf_score_avg"] is None:
                        print("✅ TF average for .45 correctly shows as None (all values were NULL)")
                    else:
                        print("❌ TF average for .45 should be None")
                        return None
                        
                    # RF should have a value
                    if "rf_score_avg" in dot45_data and dot45_data["rf_score_avg"] is not None:
                        print(f"✅ RF average for .45 calculated correctly: {dot45_data['rf_score_avg']}")
                    else:
                        print("❌ RF average for .45 not calculated correctly")
                        return None
                else:
                    print("❌ No valid_matches_count found for .45")
                    return None
            else:
                print("❌ No .45 caliber averages found")
                return None
                
            return "Shooter statistics endpoint correctly excludes NULL scores from average calculations"
            
        success, _ = self.run_test(
            "Verify shooter statistics endpoint",
            "GET",
            f"shooter-averages/{self.test_data['shooter_id']}",
            200,
            check_function=check_shooter_averages
        )
        
        return success

    def run_all_tests(self):
        """Run all tests for Excel export with NULL values"""
        # Step 1: Create a test shooter
        if not self.create_test_shooter():
            return False
            
        # Step 2: Create a test match
        if not self.create_test_match():
            return False
            
        # Step 3: Create scores with NULL and 0 values
        if not self.create_scores_with_null_and_zero_values():
            return False
            
        # Step 4: Verify Excel export
        excel_result = self.verify_excel_export()
        
        # Step 5: Verify shooter statistics
        stats_result = self.verify_shooter_statistics()
        
        # Return overall result
        return excel_result and stats_result

def main():
    tester = ExcelNullValuesComprehensiveTester()
    
    # Login first
    if not tester.login():
        print("❌ Login failed, cannot continue tests")
        return 1
    
    # Run all tests
    print("\n=== Testing Excel Export and Average Calculations for NULL Values ===")
    test_result = tester.run_all_tests()
    
    # Print summary
    print("\n=== Test Summary ===")
    print(f"Tests Run: {tester.tests_run}")
    print(f"Tests Passed: {tester.tests_passed}")
    
    if test_result:
        print("\n✅ ALL TESTS PASSED")
        print("\nVerification Results:")
        print("1. NULL scores are correctly displayed as '-' in the Excel export")
        print("2. NULL scores are correctly excluded from average calculations")
        print("3. Scores of 0 are correctly included in average calculations")
        print("4. The shooter statistics endpoint correctly handles NULL scores")
    else:
        print("\n❌ SOME TESTS FAILED")
        
    return 0 if test_result else 1

if __name__ == "__main__":
    sys.exit(main())

"""
Shared OpenPyXL styles for Match Track Excel exports.

Keep formatting consistent and conservative — polish without breaking
column layout or data values.
"""

from __future__ import annotations

from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.page import PageMargins

# --- Palette (professional, print-friendly) ---
NAVY = "1F4E79"
NAVY_LIGHT = "2E75B6"
GOLD = "FFF2CC"  # place / award highlight
GOLD_DEEP = "FFE699"
SECTION = "D6DCE4"  # section banner
HEADER_BG = "1F4E79"
HEADER_FG = "FFFFFF"
ALT_ROW = "F2F2F2"
SPECIAL = "E2EFDA"  # soft green for special awards
THIN = Border(
    left=Side(style="thin", color="B0B0B0"),
    right=Side(style="thin", color="B0B0B0"),
    top=Side(style="thin", color="B0B0B0"),
    bottom=Side(style="thin", color="B0B0B0"),
)


def font_title() -> Font:
    return Font(name="Calibri", bold=True, size=16, color=NAVY)


def font_subtitle() -> Font:
    return Font(name="Calibri", bold=True, size=12, color=NAVY)


def font_section() -> Font:
    return Font(name="Calibri", bold=True, size=11, color="000000")


def font_header() -> Font:
    return Font(name="Calibri", bold=True, size=10, color=HEADER_FG)


def font_body(bold: bool = False) -> Font:
    return Font(name="Calibri", bold=bold, size=10)


def font_meta_label() -> Font:
    return Font(name="Calibri", bold=True, size=10, color="404040")


def fill_header() -> PatternFill:
    return PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")


def fill_section() -> PatternFill:
    return PatternFill(start_color=SECTION, end_color=SECTION, fill_type="solid")


def fill_gold() -> PatternFill:
    return PatternFill(start_color=GOLD, end_color=GOLD, fill_type="solid")


def fill_special() -> PatternFill:
    return PatternFill(start_color=SPECIAL, end_color=SPECIAL, fill_type="solid")


def fill_alt() -> PatternFill:
    return PatternFill(start_color=ALT_ROW, end_color=ALT_ROW, fill_type="solid")


def align_center() -> Alignment:
    return Alignment(horizontal="center", vertical="center", wrap_text=False)


def align_left() -> Alignment:
    return Alignment(horizontal="left", vertical="center", wrap_text=False)


def align_right() -> Alignment:
    return Alignment(horizontal="right", vertical="center", wrap_text=False)


def apply_print_setup(ws, *, landscape: bool = False, fit_width: bool = True) -> None:
    """Sensible margins and print defaults for handouts."""
    ws.page_margins = PageMargins(
        left=0.5, right=0.5, top=0.6, bottom=0.6, header=0.2, footer=0.2
    )
    ws.print_options.horizontalCentered = True
    ws.page_setup.orientation = "landscape" if landscape else "portrait"
    ws.page_setup.fitToPage = fit_width
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_view.showGridLines = False


def style_range_border(ws, min_row: int, max_row: int, min_col: int, max_col: int) -> None:
    for row in ws.iter_rows(
        min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col
    ):
        for cell in row:
            cell.border = THIN


def style_header_row(ws, row: int, max_col: int) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_header()
        cell.fill = fill_header()
        cell.alignment = align_center()
        cell.border = THIN


def style_section_banner(ws, row: int, max_col: int = 5) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.fill = fill_section()
        cell.font = font_section()
        cell.border = THIN
    # Merge looks nicer for section titles when first cell has the text
    try:
        ws.merge_cells(
            start_row=row, start_column=1, end_row=row, end_column=max_col
        )
    except Exception:
        pass


def autosize_columns(ws, min_width: int = 10, max_width: int = 36) -> None:
    for col_cells in ws.columns:
        letter = get_column_letter(col_cells[0].column)
        length = 0
        for cell in col_cells:
            if cell.value is None:
                continue
            length = max(length, len(str(cell.value)))
        ws.column_dimensions[letter].width = max(
            min_width, min(max_width, length + 2)
        )


def style_data_row(
    ws,
    row: int,
    max_col: int,
    *,
    alt: bool = False,
    highlight: bool = False,
    special: bool = False,
) -> None:
    for col in range(1, max_col + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = font_body(bold=highlight)
        cell.border = THIN
        cell.alignment = align_center() if col != 3 else align_left()
        if highlight:
            cell.fill = fill_gold()
        elif special:
            cell.fill = fill_special()
        elif alt:
            cell.fill = fill_alt()

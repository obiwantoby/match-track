from fastapi import FastAPI, APIRouter, HTTPException, Body, Depends, status, UploadFile, File
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os
import logging
import uuid
import io
import csv
from typing import Dict, List, Optional, Any, Union
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from fastapi.responses import StreamingResponse
from pathlib import Path
from pydantic import BaseModel, Field, EmailStr, ValidationError
from datetime import datetime, timedelta
from enum import Enum
import re

from .core import (
    BasicMatchType,
    AggregateType,
    CaliberType,
    Rating,
    Division,
    SpecialCategory,
    STANDARD_CALIBER_ORDER_MAP,
    ShooterBase,
    Shooter,
    MatchTypeInstance,
    MatchBase,
    Match,
    LeagueBase,
    League,
    ScoreStage, # This was already here
    ScoreBase,  # This was already here
    Score,      # This was already here
    get_stages_for_match_type,      # ADD THIS
    get_match_type_max_score,        # ADD THIS
    _get_aggregate_components,        # ADD THIS
    _get_ordered_calibers_for_aggregate, # ADD THIS
    calculate_aggregates,              # ADD THIS
    calculate_shooter_averages_by_caliber,  # ADD THIS
    calculate_score_subtotals               # ADD THIS
)
from .bulletin import (
    CompetitorResult,
    build_bulletin,
    event_score_from_score_doc,
)

# Import auth components
from .auth import (
    auth_router,
    get_current_active_user,
    get_admin_user,
    User,
    UserBase,
    UserCreate,
    UserInDB,
    UserRole,
    get_password_hash,
    create_user_record,
)
from .database import db, connect_to_mongo, close_mongo_connection

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / ".env")

logging.basicConfig(
    level=logging.INFO, format="%(asctime)s - %(name)s - %(levelname)s - %(message)s"
)
logger = logging.getLogger(__name__)

# Create the main app without a prefix
app = FastAPI()

# Create a router with the /api prefix
api_router = APIRouter(prefix="/api")

# --- Helper functions that remain in server.py (Excel specific) ---
# _build_dynamic_aggregate_header_and_calibers
# _build_dynamic_non_aggregate_header
# build_aggregate_row_grouped
# build_non_aggregate_row

# --- New dynamic header builder for aggregate matches ---
def _build_dynamic_aggregate_header_and_calibers(match_obj: Match):
    header_row1 = ["", ""]  # For Shooter, Aggregate Total
    header_row2 = ["Shooter", "Aggregate Total"]
    
    base_match_type_for_agg, agg_sub_fields = _get_aggregate_components(match_obj.aggregate_type)
    
    if not base_match_type_for_agg or not agg_sub_fields:
        # Fallback for unknown aggregate type or if no components defined
        return header_row1, header_row2, [], [], None 

    ordered_calibers_list = _get_ordered_calibers_for_aggregate(match_obj, base_match_type_for_agg)

    for caliber_enum in ordered_calibers_list:
        caliber_str = caliber_enum.value
        header_row1 += [caliber_str] + [""] * (len(agg_sub_fields) - 1)
        header_row2 += agg_sub_fields
        
    return header_row1, header_row2, ordered_calibers_list, agg_sub_fields, base_match_type_for_agg

# --- New dynamic header builder for non-aggregate matches ---
def _build_dynamic_non_aggregate_header(match_obj: Match) -> List[str]:
    header = ["Shooter", "Average"]
    
    # Use the SAME order as stored in match_obj.match_types (preserve creation order)
    for mt in match_obj.match_types:
        # Use the SAME order as stored in mt.calibers (preserve creation order)
        for caliber_enum in mt.calibers:
            header.append(f"{mt.instance_name} ({caliber_enum.value})")
    return header

# --- Modified build_aggregate_row_grouped function ---
def build_aggregate_row_grouped(
    shooter: Shooter, 
    shooter_data: Dict[str, Any], 
    report_data: Dict[str, Any], 
    ordered_calibers: List[CaliberType], 
    agg_sub_fields: List[str],
    base_match_type_for_agg: BasicMatchType
):
    row = [shooter.name]
    
    overall_agg_total_score = 0
    overall_agg_total_x = 0
    
    match_config = report_data.get("match_config", {})
    match_types_configs_dict = {mtc["instance_name"]: mtc for mtc in match_config.get("match_types", [])}

    for score_key, score_item in shooter_data["scores"].items():
        score_details = score_item["score"]
        if score_details.get("not_shot", False) or score_details["total_score"] is None:
            continue

        instance_name = score_details["match_type_instance"]
        mt_config_for_score = match_types_configs_dict.get(instance_name)
        
        if mt_config_for_score and mt_config_for_score["type"] == base_match_type_for_agg:
            overall_agg_total_score += score_details["total_score"]
            overall_agg_total_x += (score_details["total_x_count"] or 0)
            
    row.append(f"{overall_agg_total_score} ({overall_agg_total_x}X)" if overall_agg_total_score > 0 or overall_agg_total_x > 0 else "-")

    for target_caliber_enum in ordered_calibers:
        target_caliber_str = target_caliber_enum.value
        
        col_data = {field: {"score": 0, "x": 0} for field in agg_sub_fields}
        col_total_score_direct_sum = 0 # Sum of total_score from individual matches for this caliber
        col_total_x_direct_sum = 0     # Sum of total_x_count from individual matches for this caliber
        has_data_for_caliber_column = False

        for score_key, score_item in shooter_data["scores"].items():
            score_details = score_item["score"]
            if score_details.get("not_shot", False) or score_details["caliber"] != target_caliber_str:
                continue

            instance_name = score_details["match_type_instance"]
            mt_config_for_score = match_types_configs_dict.get(instance_name)

            if mt_config_for_score and mt_config_for_score["type"] == base_match_type_for_agg:
                has_data_for_caliber_column = True
                
                if score_details["total_score"] is not None:
                    col_total_score_direct_sum += score_details["total_score"]
                if score_details["total_x_count"] is not None:
                    col_total_x_direct_sum += score_details["total_x_count"]

                for stage in score_details["stages"]:
                    val = stage["score"] if stage["score"] is not None else 0
                    xval = stage["x_count"] if stage["x_count"] is not None else 0
                    
                    stage_name_upper = stage["name"].upper() # Normalize stage name for matching

                    if stage_name_upper.startswith("SF") and "NMC" not in stage_name_upper:
                        col_data["SF"]["score"] += val
                        col_data["SF"]["x"] += xval
                    elif "NMC" in stage_name_upper and "NMC" in agg_sub_fields:
                         col_data["NMC"]["score"] += val
                         col_data["NMC"]["x"] += xval
                    elif stage_name_upper.startswith("TF") and "NMC" not in stage_name_upper:
                        col_data["TF"]["score"] += val
                        col_data["TF"]["x"] += xval
                    elif stage_name_upper.startswith("RF") and "NMC" not in stage_name_upper:
                        col_data["RF"]["score"] += val
                        col_data["RF"]["x"] += xval
        
        def fmt_cell(score_val, x_val):
            if not has_data_for_caliber_column: return "-"
            return f"{score_val} ({x_val}X)" if score_val > 0 or x_val > 0 else "0 (0X)"

        for field_name in agg_sub_fields:
            if field_name in ["SF", "TF", "RF"] or (field_name == "NMC" and "NMC" in col_data):
                row.append(fmt_cell(col_data[field_name]["score"], col_data[field_name]["x"]))
            elif field_name == "900" or field_name == "600": # This is the total for the caliber column
                row.append(fmt_cell(col_total_score_direct_sum, col_total_x_direct_sum))
            # else: # Should not happen with defined agg_sub_fields
            #     row.append("-") # Fallback for unexpected field
                
    return row

# --- New function to build a row for non-aggregate matches ---
def build_non_aggregate_row(
    shooter: Shooter, 
    shooter_data: Dict[str, Any], 
    match_obj: Match # Pass the full match_obj to access its structure
) -> List[Any]:
    row = [shooter.name]
    
    total_score_sum = 0
    num_scored_entries = 0
    
    # Pre-calculate overall average for this shooter in this non-aggregate match
    for score_key, score_item in shooter_data["scores"].items():
        score_details = score_item["score"]
        if not score_details.get("not_shot", False) and score_details["total_score"] is not None:
            total_score_sum += score_details["total_score"]
            num_scored_entries += 1
            
    if num_scored_entries > 0:
        average_score = round(total_score_sum / num_scored_entries, 2)
        row.append(average_score)
    else:
        row.append("-") # No scores to average

    # Use the SAME order as in header builder - preserve match creation order
    for mt_instance in match_obj.match_types:
        # Use the SAME order as in header builder - preserve caliber creation order
        for target_caliber_enum in mt_instance.calibers:
            target_caliber_str = target_caliber_enum.value
            
            score_found = False
            # Iterate through the shooter's scores to find the matching entry
            for score_key, score_item in shooter_data["scores"].items():
                score_details = score_item["score"]
                # The key in shooter_data["scores"] is typically f"{match_type_instance}_{caliber}"
                # We need to match instance_name and caliber
                if score_details["match_type_instance"] == mt_instance.instance_name and \
                   score_details["caliber"] == target_caliber_str:
                    if not score_details.get("not_shot", False) and score_details["total_score"] is not None:
                        row.append(f"{score_details['total_score']} ({score_details.get('total_x_count', 0)}X)")
                    else:
                        row.append("-") # Not shot or no score
                    score_found = True
                    break
            
            if not score_found:
                row.append("-") # No score entry found for this specific match_type_instance and caliber
                
    return row

# Define Models
class ShooterCreate(ShooterBase):
    pass


class LeagueCreate(LeagueBase):
    pass


class LeagueUpdate(LeagueBase):
    pass


class MatchCreate(MatchBase):
    """Match structure + optional league link used only at create time for seeding."""
    league_id: Optional[str] = None


class MatchLeagueLink(BaseModel):
    """Attach or detach a match from a league without touching structure/scores."""
    league_id: Optional[str] = None
    # When attaching, also pull any league members missing from the match roster
    pull_roster: bool = True


# Score Stage (individual component scores)
class ScoreStage(BaseModel):
    name: str  # e.g., "SF", "TF1", "RF2"
    score: Optional[int] = None
    x_count: Optional[int] = None


# Score Entry
class ScoreBase(BaseModel):
    shooter_id: str
    match_id: str
    caliber: CaliberType
    match_type_instance: str  # e.g., "NMC1", "600_1"
    stages: List[ScoreStage]
    total_score: Optional[int] = None
    total_x_count: Optional[int] = None
    not_shot: bool = False  # New field to indicate if the match was not shot


class ScoreCreate(ScoreBase):
    pass


class Score(ScoreBase):
    id: str = Field(default_factory=lambda: str(uuid.uuid4()))
    created_at: datetime = Field(default_factory=datetime.utcnow)


class ScoreWithDetails(Score):
    shooter_name: str
    match_name: str
    match_date: datetime
    match_location: str


# --- User bulk-import models ---
class BulkUserRowResult(BaseModel):
    row: int
    email: Optional[str] = None
    username: Optional[str] = None
    status: str  # created | skipped | error
    detail: Optional[str] = None


class BulkUserImportResult(BaseModel):
    created: int
    skipped: int
    errors: int
    results: List[BulkUserRowResult]


def _normalize_csv_headers(headers: List[str]) -> Dict[str, str]:
    """Map lowercase/stripped header names to original header keys."""
    mapping = {}
    for h in headers:
        if h is None:
            continue
        key = str(h).strip().lower().replace(" ", "_")
        mapping[key] = h
    return mapping


def _parse_role(value: Optional[str]) -> UserRole:
    if value is None or str(value).strip() == "":
        return UserRole.REPORTER
    normalized = str(value).strip().lower()
    if normalized in ("admin", "administrator"):
        return UserRole.ADMIN
    if normalized in ("reporter", "user"):
        return UserRole.REPORTER
    raise ValueError(f"Invalid role '{value}'. Use 'admin' or 'reporter'.")


# User management routes (admin only)
@api_router.get("/users", response_model=List[User])
async def get_users(current_user: User = Depends(get_admin_user)):
    users = await db.users.find().to_list(1000)
    return [User(**user) for user in users]


@api_router.post("/users", response_model=User, status_code=status.HTTP_201_CREATED)
async def create_user(
    user_data: UserCreate, current_user: User = Depends(get_admin_user)
):
    """Admin-only: create a single user with an explicit role."""
    try:
        return await create_user_record(
            email=user_data.email,
            username=user_data.username,
            password=user_data.password,
            role=user_data.role,
        )
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Admin user create error: {str(e)}")
        raise HTTPException(
            status_code=status.HTTP_500_INTERNAL_SERVER_ERROR,
            detail=f"Failed to create user: {str(e)}",
        )


@api_router.post("/users/bulk-csv", response_model=BulkUserImportResult)
async def bulk_create_users_from_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_user),
):
    """
    Admin-only: bulk-create users from a CSV upload.

    Required columns: username, email, password
    Optional columns: role (admin|reporter, default reporter)

    Header names are case-insensitive. Extra columns are ignored.
    Existing emails are skipped (not overwritten).
    """
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload a .csv file",
        )

    raw = await file.read()
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty",
        )

    # Handle UTF-8 BOM from Excel exports
    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except UnicodeDecodeError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Could not decode CSV file: {e}",
            )

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV has no header row",
        )

    header_map = _normalize_csv_headers(list(reader.fieldnames))
    required = ("username", "email", "password")
    missing = [col for col in required if col not in header_map]
    if missing:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"CSV is missing required column(s): {', '.join(missing)}. "
                "Expected headers: username, email, password[, role]"
            ),
        )

    results: List[BulkUserRowResult] = []
    created = skipped = errors = 0

    def cell(row: dict, col: str) -> str:
        original = header_map.get(col)
        if original is None:
            return ""
        val = row.get(original)
        return "" if val is None else str(val).strip()

    for row_num, row in enumerate(reader, start=2):  # row 1 is header
        username = cell(row, "username")
        email = cell(row, "email")
        password = cell(row, "password")
        role_raw = cell(row, "role") if "role" in header_map else ""

        if not username and not email and not password:
            # Skip blank lines
            continue

        if not username or not email or not password:
            errors += 1
            results.append(
                BulkUserRowResult(
                    row=row_num,
                    email=email or None,
                    username=username or None,
                    status="error",
                    detail="username, email, and password are required",
                )
            )
            continue

        try:
            role = _parse_role(role_raw)
            # Validate email format via UserCreate
            UserCreate(email=email, username=username, password=password, role=role)
            user = await create_user_record(
                email=email,
                username=username,
                password=password,
                role=role,
            )
            created += 1
            results.append(
                BulkUserRowResult(
                    row=row_num,
                    email=user.email,
                    username=user.username,
                    status="created",
                    detail=f"Created with role {user.role.value}",
                )
            )
        except HTTPException as he:
            # Email already registered → skip rather than fail the whole batch
            if he.status_code == status.HTTP_400_BAD_REQUEST and "already" in str(
                he.detail
            ).lower():
                skipped += 1
                results.append(
                    BulkUserRowResult(
                        row=row_num,
                        email=email,
                        username=username,
                        status="skipped",
                        detail=str(he.detail),
                    )
                )
            else:
                errors += 1
                results.append(
                    BulkUserRowResult(
                        row=row_num,
                        email=email,
                        username=username,
                        status="error",
                        detail=str(he.detail),
                    )
                )
        except (ValidationError, ValueError) as e:
            errors += 1
            detail = str(e)
            if isinstance(e, ValidationError):
                # Compact pydantic errors
                detail = "; ".join(
                    f"{'.'.join(str(x) for x in err.get('loc', ()))}: {err.get('msg')}"
                    for err in e.errors()
                )
            results.append(
                BulkUserRowResult(
                    row=row_num,
                    email=email,
                    username=username,
                    status="error",
                    detail=detail,
                )
            )
        except Exception as e:
            errors += 1
            logger.error(f"CSV row {row_num} import error: {e}")
            results.append(
                BulkUserRowResult(
                    row=row_num,
                    email=email,
                    username=username,
                    status="error",
                    detail=f"Unexpected error: {e}",
                )
            )

    if created == 0 and skipped == 0 and errors == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV contained no data rows",
        )

    return BulkUserImportResult(
        created=created,
        skipped=skipped,
        errors=errors,
        results=results,
    )


@api_router.put("/users/{user_id}", response_model=User)
async def update_user(
    user_id: str, user_data: UserBase, current_user: User = Depends(get_admin_user)
):
    # Check if user exists
    existing_user = await db.users.find_one({"id": user_id})
    if not existing_user:
        raise HTTPException(status_code=404, detail="User not found")

    # Prevent demoting/changing your own role out of admin accidentally
    if user_id == current_user.id and user_data.role != UserRole.ADMIN:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot change your own role away from admin",
        )

    await db.users.update_one({"id": user_id}, {"$set": user_data.dict()})

    updated_user = await db.users.find_one({"id": user_id})
    return User(**updated_user)


@api_router.delete("/users/{user_id}", response_model=Dict[str, bool])
async def delete_user(user_id: str, current_user: User = Depends(get_admin_user)):
    # Prevent deleting the current user
    if user_id == current_user.id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Cannot delete your own account",
        )

    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(status_code=404, detail="User not found")

    return {"success": True}


# Database reset endpoint (admin only)
@api_router.post("/reset-database", response_model=Dict[str, bool])
async def reset_database(current_user: User = Depends(get_admin_user)):
    """Reset the database to a clean state, preserving only the current admin user"""
    # Get current admin user details
    admin_user = await db.users.find_one({"id": current_user.id})

    # Drop all collections
    for collection_name in await db.list_collection_names():
        if collection_name != "users":  # Don't drop users yet
            await db[collection_name].drop()

    # Remove all users except the current admin
    await db.users.delete_many({"id": {"$ne": current_user.id}})

    # Return success
    return {"success": True}


# --- Shooter bulk-import models ---
class BulkShooterRowResult(BaseModel):
    row: int
    name: Optional[str] = None
    status: str  # created | skipped | error
    detail: Optional[str] = None


class BulkShooterImportResult(BaseModel):
    created: int
    skipped: int
    errors: int
    results: List[BulkShooterRowResult]


def _empty_to_none(value: Optional[str]) -> Optional[str]:
    if value is None:
        return None
    stripped = str(value).strip()
    return stripped if stripped else None


def _parse_rating(value: Optional[str]) -> Optional[Rating]:
    cleaned = _empty_to_none(value)
    if cleaned is None:
        return None
    normalized = cleaned.upper()
    # Allow common long-form aliases
    aliases = {
        "HIGH MASTER": "HM",
        "HIGHMASTER": "HM",
        "MASTER": "MA",
        "EXPERT": "EX",
        "SHARPSHOOTER": "SS",
        "MARKSMAN": "MK",
        "UNCLASSIFIED": "UNC",
        "UNCLASS": "UNC",
    }
    normalized = aliases.get(normalized, normalized)
    try:
        return Rating(normalized)
    except ValueError:
        valid = ", ".join(r.value for r in Rating)
        raise ValueError(f"Invalid rating '{value}'. Use one of: {valid}")


async def _create_shooter_record(
    name: str,
    nra_number: Optional[str] = None,
    cmp_number: Optional[str] = None,
    rating: Optional[Rating] = None,
    competitor_number: Optional[int] = None,
    division: Optional[Division] = Division.CIVILIAN,
    special_categories: Optional[List] = None,
    *,
    skip_if_duplicate: bool = False,
) -> tuple[Optional[Shooter], Optional[str]]:
    """
    Insert a shooter. Returns (shooter, skip_reason).
    If skip_if_duplicate and a same-name shooter exists, returns (None, reason).
    """
    name = name.strip()
    if not name:
        raise ValueError("name is required")

    if skip_if_duplicate:
        existing = await db.shooters.find_one(
            {"name": {"$regex": f"^{re.escape(name)}$", "$options": "i"}}
        )
        if existing:
            return None, f"Shooter named '{existing['name']}' already exists"

        # Also skip when NRA number collides (when provided)
        if nra_number:
            by_nra = await db.shooters.find_one({"nra_number": nra_number})
            if by_nra:
                return None, f"NRA number {nra_number} already used by '{by_nra['name']}'"

    shooter_obj = Shooter(
        name=name,
        nra_number=nra_number,
        cmp_number=cmp_number,
        rating=rating,
        competitor_number=competitor_number,
        division=division or Division.CIVILIAN,
        special_categories=list(special_categories or []),
    )
    await db.shooters.insert_one(shooter_obj.dict())
    return shooter_obj, None


# Shooter Routes
@api_router.post("/shooters", response_model=Shooter)
async def create_shooter(
    shooter: ShooterCreate, current_user: User = Depends(get_current_active_user)
):
    # Only admins can create shooters
    if current_user.role == UserRole.REPORTER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions"
        )

    # Normalize blank optional fields
    data = shooter.dict()
    data["nra_number"] = _empty_to_none(data.get("nra_number"))
    data["cmp_number"] = _empty_to_none(data.get("cmp_number"))
    if data.get("rating") == "" or data.get("rating") is None:
        data["rating"] = None

    shooter_obj, _ = await _create_shooter_record(
        name=data["name"],
        nra_number=data.get("nra_number"),
        cmp_number=data.get("cmp_number"),
        rating=data.get("rating"),
        competitor_number=data.get("competitor_number"),
        division=data.get("division"),
        special_categories=data.get("special_categories"),
        skip_if_duplicate=False,
    )
    return shooter_obj


@api_router.post("/shooters/bulk-csv", response_model=BulkShooterImportResult)
async def bulk_create_shooters_from_csv(
    file: UploadFile = File(...),
    current_user: User = Depends(get_admin_user),
):
    """
    Admin-only: bulk-create shooter profiles from a CSV upload.

    Required columns: name
    Optional columns: nra_number, cmp_number, rating

    Header names are case-insensitive (spaces/underscores ok).
    Duplicate names (case-insensitive) and duplicate NRA numbers are skipped.
    """
    if not file.filename or not file.filename.lower().endswith(".csv"):
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Please upload a .csv file",
        )

    raw = await file.read()
    if not raw:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Uploaded file is empty",
        )

    try:
        text = raw.decode("utf-8-sig")
    except UnicodeDecodeError:
        try:
            text = raw.decode("latin-1")
        except UnicodeDecodeError as e:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Could not decode CSV file: {e}",
            )

    reader = csv.DictReader(io.StringIO(text))
    if not reader.fieldnames:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV has no header row",
        )

    header_map = _normalize_csv_headers(list(reader.fieldnames))
    # Accept common aliases for the name column
    if "name" not in header_map:
        for alias in ("shooter", "shooter_name", "full_name", "competitor"):
            if alias in header_map:
                header_map["name"] = header_map[alias]
                break

    if "name" not in header_map:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                "CSV is missing required column: name. "
                "Expected headers: name[, nra_number, cmp_number, rating]"
            ),
        )

    # Alias optional columns
    if "nra_number" not in header_map:
        for alias in ("nra", "nra_num", "nra#"):
            if alias in header_map:
                header_map["nra_number"] = header_map[alias]
                break
    if "cmp_number" not in header_map:
        for alias in ("cmp", "cmp_num", "cmp#"):
            if alias in header_map:
                header_map["cmp_number"] = header_map[alias]
                break
    if "special_categories" not in header_map:
        for alias in ("specials", "special_category", "categories"):
            if alias in header_map:
                header_map["special_categories"] = header_map[alias]
                break
    if "division" not in header_map:
        for alias in ("div", "category_division"):
            if alias in header_map:
                header_map["division"] = header_map[alias]
                break
    if "competitor_number" not in header_map:
        for alias in ("comp_number", "competitor_no", "competitor#", "number"):
            if alias in header_map:
                header_map["competitor_number"] = header_map[alias]
                break

    def cell(row: dict, col: str) -> str:
        original = header_map.get(col)
        if original is None:
            return ""
        val = row.get(original)
        return "" if val is None else str(val).strip()

    def parse_specials(raw: str) -> List[str]:
        if not raw:
            return []
        parts = [p.strip() for p in re.split(r"[|;,/]", raw) if p.strip()]
        aliases = {
            "grand senior": "Grand Senior",
            "gs": "Grand Senior",
            "senior": "Senior",
            "women": "Women",
            "woman": "Women",
            "veteran": "Veteran",
            "vet": "Veteran",
        }
        out: List[str] = []
        for p in parts:
            canon = aliases.get(p.casefold())
            if canon:
                out.append(canon)
        seen: set = set()
        uniq: List[str] = []
        for x in out:
            if x not in seen:
                seen.add(x)
                uniq.append(x)
        return uniq

    def parse_division(raw: str) -> Optional[str]:
        if not raw:
            return "Civilian"
        key = raw.strip().casefold()
        mapping = {
            "civilian": "Civilian",
            "civ": "Civilian",
            "police": "Police",
            "service": "Service",
            "police/service": "Police",
        }
        return mapping.get(key, raw if raw in ("Civilian", "Police", "Service") else "Civilian")

    results: List[BulkShooterRowResult] = []
    created = skipped = errors = 0
    # Track names/nra seen in this file to avoid double-insert within same upload
    seen_names: set[str] = set()
    seen_nra: set[str] = set()

    for row_num, row in enumerate(reader, start=2):
        name = cell(row, "name")
        nra_number = _empty_to_none(cell(row, "nra_number"))
        cmp_number = _empty_to_none(cell(row, "cmp_number"))
        rating_raw = cell(row, "rating") if "rating" in header_map else ""
        specials_raw = cell(row, "special_categories") if "special_categories" in header_map else ""
        division_raw = cell(row, "division") if "division" in header_map else ""
        comp_raw = cell(row, "competitor_number") if "competitor_number" in header_map else ""

        if not name and not nra_number and not cmp_number and not rating_raw:
            continue

        if not name:
            errors += 1
            results.append(
                BulkShooterRowResult(
                    row=row_num,
                    name=None,
                    status="error",
                    detail="name is required",
                )
            )
            continue

        name_key = name.casefold()
        if name_key in seen_names:
            skipped += 1
            results.append(
                BulkShooterRowResult(
                    row=row_num,
                    name=name,
                    status="skipped",
                    detail="Duplicate name in this CSV",
                )
            )
            continue

        if nra_number and nra_number in seen_nra:
            skipped += 1
            results.append(
                BulkShooterRowResult(
                    row=row_num,
                    name=name,
                    status="skipped",
                    detail=f"Duplicate NRA number {nra_number} in this CSV",
                )
            )
            continue

        try:
            rating = _parse_rating(rating_raw)
            specials = parse_specials(specials_raw)
            division = parse_division(division_raw)
            competitor_number = None
            if comp_raw:
                try:
                    competitor_number = int(float(comp_raw))
                except ValueError:
                    competitor_number = None
            shooter_obj, skip_reason = await _create_shooter_record(
                name=name,
                nra_number=nra_number,
                cmp_number=cmp_number,
                rating=rating,
                competitor_number=competitor_number,
                division=Division(division) if division else Division.CIVILIAN,
                special_categories=[SpecialCategory(c) for c in specials],
                skip_if_duplicate=True,
            )
            if skip_reason:
                skipped += 1
                results.append(
                    BulkShooterRowResult(
                        row=row_num,
                        name=name,
                        status="skipped",
                        detail=skip_reason,
                    )
                )
            else:
                created += 1
                seen_names.add(name_key)
                if nra_number:
                    seen_nra.add(nra_number)
                results.append(
                    BulkShooterRowResult(
                        row=row_num,
                        name=shooter_obj.name,
                        status="created",
                        detail="Created",
                    )
                )
        except (ValidationError, ValueError) as e:
            errors += 1
            detail = str(e)
            if isinstance(e, ValidationError):
                detail = "; ".join(
                    f"{'.'.join(str(x) for x in err.get('loc', ()))}: {err.get('msg')}"
                    for err in e.errors()
                )
            results.append(
                BulkShooterRowResult(
                    row=row_num,
                    name=name,
                    status="error",
                    detail=detail,
                )
            )
        except Exception as e:
            errors += 1
            logger.error(f"Shooter CSV row {row_num} import error: {e}")
            results.append(
                BulkShooterRowResult(
                    row=row_num,
                    name=name,
                    status="error",
                    detail=f"Unexpected error: {e}",
                )
            )

    if created == 0 and skipped == 0 and errors == 0:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="CSV contained no data rows",
        )

    return BulkShooterImportResult(
        created=created,
        skipped=skipped,
        errors=errors,
        results=results,
    )


@api_router.get("/shooters", response_model=List[Shooter])
async def get_shooters(current_user: User = Depends(get_current_active_user)):
    shooters = await db.shooters.find().to_list(1000)
    # Stable alphabetical order for dropdowns / management
    parsed = []
    for shooter in shooters:
        shooter.setdefault("division", "Civilian")
        shooter.setdefault("special_categories", [])
        try:
            parsed.append(Shooter(**shooter))
        except Exception as e:
            logger.warning(f"Skipping invalid shooter doc: {e}")
    parsed.sort(key=lambda s: (s.name or "").casefold())
    return parsed


@api_router.get("/shooters/{shooter_id}", response_model=Shooter)
async def get_shooter(
    shooter_id: str, current_user: User = Depends(get_current_active_user)
):
    shooter = await db.shooters.find_one({"id": shooter_id})
    if not shooter:
        raise HTTPException(status_code=404, detail="Shooter not found")
    shooter.setdefault("division", "Civilian")
    shooter.setdefault("special_categories", [])
    return Shooter(**shooter)


@api_router.put("/shooters/{shooter_id}", response_model=Shooter)
async def update_shooter(
    shooter_id: str,
    shooter_update: ShooterCreate,
    current_user: User = Depends(get_admin_user),
):
    """Admin-only: update shooter profile fields. Does not touch scores."""
    existing = await db.shooters.find_one({"id": shooter_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Shooter not found")

    data = shooter_update.dict()
    name = (data.get("name") or "").strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="name is required"
        )

    nra_number = _empty_to_none(data.get("nra_number"))
    cmp_number = _empty_to_none(data.get("cmp_number"))
    rating = data.get("rating")
    if rating == "" or rating is None:
        rating = None
    competitor_number = data.get("competitor_number")
    division = data.get("division") or Division.CIVILIAN
    special_categories = data.get("special_categories") or []

    # Prevent NRA collisions with a *different* shooter (same id is always allowed)
    if nra_number:
        by_nra = await db.shooters.find_one(
            {"nra_number": str(nra_number), "id": {"$ne": shooter_id}}
        )
        if by_nra:
            other = by_nra.get("name") or by_nra.get("id")
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=(
                    f"NRA number {nra_number} already used by '{other}'. "
                    "That is a separate shooter record (often a seed duplicate with a "
                    "different name spelling). Edit that shooter, clear one NRA number, "
                    "or delete the duplicate."
                ),
            )

    update_fields = {
        "name": name,
        "nra_number": nra_number,
        "cmp_number": cmp_number,
        "rating": rating.value if isinstance(rating, Rating) else rating,
        "competitor_number": competitor_number,
        "division": division.value if isinstance(division, Division) else division,
        "special_categories": [
            c.value if hasattr(c, "value") else c for c in special_categories
        ],
    }
    await db.shooters.update_one({"id": shooter_id}, {"$set": update_fields})
    updated = await db.shooters.find_one({"id": shooter_id})
    # Defaults for older documents
    updated.setdefault("division", "Civilian")
    updated.setdefault("special_categories", [])
    return Shooter(**updated)


@api_router.delete("/shooters/{shooter_id}")
async def delete_shooter(
    shooter_id: str,
    force: bool = False,
    current_user: User = Depends(get_admin_user),
):
    """
    Admin-only: delete a shooter profile.

    By default refuses if the shooter has any scores. Pass force=true to also
    delete all of that shooter's scores (across all matches). Also removes the
    shooter from every match roster.
    """
    existing = await db.shooters.find_one({"id": shooter_id})
    if not existing:
        raise HTTPException(status_code=404, detail="Shooter not found")

    score_count = await db.scores.count_documents({"shooter_id": shooter_id})
    if score_count > 0 and not force:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Shooter has {score_count} score(s). "
                "Pass force=true to delete the shooter and all their scores, "
                "or remove scores first."
            ),
        )

    deleted_scores = 0
    if score_count > 0 and force:
        result = await db.scores.delete_many({"shooter_id": shooter_id})
        deleted_scores = result.deleted_count

    # Remove from all match rosters (safe even if field missing)
    await db.matches.update_many(
        {}, {"$pull": {"roster_shooter_ids": shooter_id}}
    )

    await db.shooters.delete_one({"id": shooter_id})
    return {
        "success": True,
        "deleted_scores": deleted_scores,
        "message": f"Deleted shooter '{existing.get('name')}'"
        + (f" and {deleted_scores} score(s)" if deleted_scores else ""),
    }


# --- Match roster models ---
class MatchRosterMember(BaseModel):
    shooter: Shooter
    score_count: int = 0
    has_scores: bool = False


class MatchRosterResponse(BaseModel):
    match_id: str
    members: List[MatchRosterMember]
    # Shooters with scores in this match who are not on the formal roster
    scored_but_not_on_roster: List[MatchRosterMember] = []


class MatchRosterAddRequest(BaseModel):
    """Add existing shooters and/or create new ones onto a match roster."""
    shooter_ids: List[str] = Field(default_factory=list)
    new_shooters: List[ShooterCreate] = Field(default_factory=list)


# --- League Routes ---
# Layering:
#   Shooter  = global person (persists forever)
#   League   = evolving roster for a club/series/season
#   Match    = one event; roster is a snapshot (can diverge: guests, no-shows)


@api_router.get("/leagues", response_model=List[League])
async def list_leagues(current_user: User = Depends(get_current_active_user)):
    docs = await db.leagues.find().sort("name", 1).to_list(1000)
    return [League(**d) for d in docs]


@api_router.post("/leagues", response_model=League, status_code=status.HTTP_201_CREATED)
async def create_league(
    body: LeagueCreate, current_user: User = Depends(get_admin_user)
):
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="League name is required"
        )
    league = League(
        name=name,
        season=_empty_to_none(body.season),
        description=_empty_to_none(body.description),
    )
    await db.leagues.insert_one(league.dict())
    return league


@api_router.get("/leagues/{league_id}", response_model=League)
async def get_league(
    league_id: str, current_user: User = Depends(get_current_active_user)
):
    doc = await db.leagues.find_one({"id": league_id})
    if not doc:
        raise HTTPException(status_code=404, detail="League not found")
    return League(**doc)


@api_router.put("/leagues/{league_id}", response_model=League)
async def update_league(
    league_id: str,
    body: LeagueUpdate,
    current_user: User = Depends(get_admin_user),
):
    existing = await db.leagues.find_one({"id": league_id})
    if not existing:
        raise HTTPException(status_code=404, detail="League not found")
    name = (body.name or "").strip()
    if not name:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST, detail="League name is required"
        )
    # Never touch roster via this endpoint
    await db.leagues.update_one(
        {"id": league_id},
        {
            "$set": {
                "name": name,
                "season": _empty_to_none(body.season),
                "description": _empty_to_none(body.description),
            }
        },
    )
    updated = await db.leagues.find_one({"id": league_id})
    return League(**updated)


@api_router.delete("/leagues/{league_id}")
async def delete_league(
    league_id: str, current_user: User = Depends(get_admin_user)
):
    """
    Delete the league record only.
    Does not delete shooters, matches, or scores.
    Unlinks matches that pointed at this league (rosters stay as-is).
    """
    existing = await db.leagues.find_one({"id": league_id})
    if not existing:
        raise HTTPException(status_code=404, detail="League not found")

    await db.matches.update_many(
        {"league_id": league_id}, {"$set": {"league_id": None}}
    )
    await db.leagues.delete_one({"id": league_id})
    return {
        "success": True,
        "message": f"Deleted league '{existing.get('name')}' and unlinked matches",
    }


class LeagueRosterResponse(BaseModel):
    league_id: str
    league_name: str
    members: List[Shooter]
    match_count: int = 0


@api_router.get("/leagues/{league_id}/roster", response_model=LeagueRosterResponse)
async def get_league_roster(
    league_id: str, current_user: User = Depends(get_current_active_user)
):
    league = await db.leagues.find_one({"id": league_id})
    if not league:
        raise HTTPException(status_code=404, detail="League not found")

    members: List[Shooter] = []
    for sid in league.get("roster_shooter_ids") or []:
        doc = await db.shooters.find_one({"id": sid})
        if doc:
            members.append(Shooter(**doc))
    members.sort(key=lambda s: (s.name or "").casefold())

    match_count = await db.matches.count_documents({"league_id": league_id})
    return LeagueRosterResponse(
        league_id=league_id,
        league_name=league.get("name") or "",
        members=members,
        match_count=match_count,
    )


@api_router.post("/leagues/{league_id}/roster", response_model=LeagueRosterResponse)
async def add_to_league_roster(
    league_id: str,
    body: MatchRosterAddRequest,
    current_user: User = Depends(get_admin_user),
):
    """Add existing and/or new shooters to the league's evolving roster."""
    league = await db.leagues.find_one({"id": league_id})
    if not league:
        raise HTTPException(status_code=404, detail="League not found")

    if not body.shooter_ids and not body.new_shooters:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Provide shooter_ids and/or new_shooters",
        )

    ids_to_add: List[str] = []
    for sid in body.shooter_ids:
        if not sid:
            continue
        exists = await db.shooters.find_one({"id": sid})
        if not exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Shooter id not found: {sid}",
            )
        ids_to_add.append(sid)

    for ns in body.new_shooters:
        data = ns.dict()
        name = (data.get("name") or "").strip()
        if not name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Each new shooter requires a name",
            )
        shooter_obj, _ = await _create_shooter_record(
            name=name,
            nra_number=_empty_to_none(data.get("nra_number")),
            cmp_number=_empty_to_none(data.get("cmp_number")),
            rating=data.get("rating") or None,
            skip_if_duplicate=False,
        )
        ids_to_add.append(shooter_obj.id)

    if ids_to_add:
        await db.leagues.update_one(
            {"id": league_id},
            {"$addToSet": {"roster_shooter_ids": {"$each": ids_to_add}}},
        )

    return await get_league_roster(league_id, current_user)


@api_router.delete("/leagues/{league_id}/roster/{shooter_id}")
async def remove_from_league_roster(
    league_id: str,
    shooter_id: str,
    current_user: User = Depends(get_admin_user),
):
    """
    Remove from league roster only.
    Does not delete the shooter, match rosters, or any scores.
    """
    league = await db.leagues.find_one({"id": league_id})
    if not league:
        raise HTTPException(status_code=404, detail="League not found")

    await db.leagues.update_one(
        {"id": league_id},
        {"$pull": {"roster_shooter_ids": shooter_id}},
    )
    return {
        "success": True,
        "message": "Removed from league roster (match rosters and scores unchanged)",
    }


# Match Routes
@api_router.post("/matches", response_model=Match)
async def create_match(
    match: MatchCreate, current_user: User = Depends(get_current_active_user)
):
    # Only admins can create matches
    if current_user.role == UserRole.REPORTER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions"
        )

    data = match.dict()
    league_id = data.pop("league_id", None) or None
    roster: List[str] = []

    if league_id:
        league = await db.leagues.find_one({"id": league_id})
        if not league:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"League not found: {league_id}",
            )
        # Seed match roster as a snapshot of the current league roster
        roster = list(league.get("roster_shooter_ids") or [])

    match_obj = Match(**data, league_id=league_id, roster_shooter_ids=roster)
    await db.matches.insert_one(match_obj.dict())
    return match_obj


@api_router.get("/matches", response_model=List[Match])
async def get_matches(current_user: User = Depends(get_current_active_user)):
    matches = await db.matches.find().sort("date", -1).to_list(1000)
    return [Match(**match) for match in matches]


@api_router.get("/matches/{match_id}", response_model=Match)
async def get_match(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")
    return Match(**match)


@api_router.put("/matches/{match_id}", response_model=Match)
async def update_match(
    match_id: str,
    match_update: MatchCreate,
    current_user: User = Depends(get_current_active_user)
):
    # Only admins can update matches
    if current_user.role != "admin":
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions"
        )
    
    # Check if match exists
    existing_match = await db.matches.find_one({"id": match_id})
    if not existing_match:
        raise HTTPException(status_code=404, detail="Match not found")
    
    existing_match_obj = Match(**existing_match)
    
    # Find match types that have been removed or modified
    existing_match_types = {mt.instance_name: mt for mt in existing_match_obj.match_types}
    new_match_types = {mt.instance_name: mt for mt in match_update.match_types}
    
    # Delete scores for entirely removed match types
    removed_match_types = set(existing_match_types.keys()) - set(new_match_types.keys())
    if removed_match_types:
        for match_type in removed_match_types:
            await db.scores.delete_many({
                "match_id": match_id,
                "match_type_instance": match_type
            })
    
    # For match types that still exist, check for removed calibers
    for match_type_name in set(existing_match_types.keys()) & set(new_match_types.keys()):
        existing_calibers = set(existing_match_types[match_type_name].calibers)
        new_calibers = set(new_match_types[match_type_name].calibers)
        
        # Find calibers that have been removed
        removed_calibers = existing_calibers - new_calibers
        
        # Delete scores for removed calibers
        if removed_calibers:
            for caliber in removed_calibers:
                await db.scores.delete_many({
                    "match_id": match_id,
                    "match_type_instance": match_type_name,
                    "caliber": caliber
                })
    
    # Structure only — never wipe roster or league link from this endpoint
    update_data = match_update.dict()
    update_data.pop("league_id", None)
    preserved_roster = list(existing_match.get("roster_shooter_ids") or [])
    preserved_league = existing_match.get("league_id")
    match_obj = Match(
        id=match_id,
        **update_data,
        roster_shooter_ids=preserved_roster,
        league_id=preserved_league,
    )
    
    # Update in database
    await db.matches.update_one(
        {"id": match_id}, 
        {"$set": match_obj.dict(exclude={"id"})}
    )
    
    # Get updated match
    updated_match = await db.matches.find_one({"id": match_id})
    if not updated_match:
        raise HTTPException(status_code=500, detail="Failed to update match")
    
    return Match(**updated_match)


@api_router.put("/matches/{match_id}/league", response_model=Match)
async def set_match_league(
    match_id: str,
    body: MatchLeagueLink,
    current_user: User = Depends(get_admin_user),
):
    """
    Link or unlink a match to a league.
    Optionally pulls missing league members onto the match roster (additive only).
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    league_id = body.league_id
    pull_ids: List[str] = []

    if league_id:
        league = await db.leagues.find_one({"id": league_id})
        if not league:
            raise HTTPException(status_code=404, detail="League not found")
        if body.pull_roster:
            pull_ids = list(league.get("roster_shooter_ids") or [])

    update: Dict[str, Any] = {"league_id": league_id}
    if pull_ids:
        await db.matches.update_one(
            {"id": match_id},
            {
                "$set": {"league_id": league_id},
                "$addToSet": {"roster_shooter_ids": {"$each": pull_ids}},
            },
        )
    else:
        await db.matches.update_one({"id": match_id}, {"$set": update})

    updated = await db.matches.find_one({"id": match_id})
    return Match(**updated)


@api_router.post(
    "/matches/{match_id}/roster/sync-from-league",
    response_model=MatchRosterResponse,
)
async def sync_match_roster_from_league(
    match_id: str, current_user: User = Depends(get_admin_user)
):
    """
    Additive sync: add any league members who are not yet on the match roster.
    Never removes match-only guests. League must be linked on the match.
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    league_id = match.get("league_id")
    if not league_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Match is not linked to a league. Link a league first.",
        )

    league = await db.leagues.find_one({"id": league_id})
    if not league:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Linked league no longer exists",
        )

    league_ids = list(league.get("roster_shooter_ids") or [])
    if league_ids:
        await db.matches.update_one(
            {"id": match_id},
            {"$addToSet": {"roster_shooter_ids": {"$each": league_ids}}},
        )

    return await get_match_roster(match_id, current_user)


@api_router.post("/matches/{match_id}/roster/{shooter_id}/promote-to-league")
async def promote_match_shooter_to_league(
    match_id: str,
    shooter_id: str,
    current_user: User = Depends(get_admin_user),
):
    """
    Grow the league over time: add a match roster shooter onto the linked league.
    Does not change other matches' rosters (they sync when you choose).
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    league_id = match.get("league_id")
    if not league_id:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Match is not linked to a league",
        )

    shooter = await db.shooters.find_one({"id": shooter_id})
    if not shooter:
        raise HTTPException(status_code=404, detail="Shooter not found")

    league = await db.leagues.find_one({"id": league_id})
    if not league:
        raise HTTPException(status_code=404, detail="League not found")

    # Ensure on match roster too (harmless if already there)
    await db.matches.update_one(
        {"id": match_id},
        {"$addToSet": {"roster_shooter_ids": shooter_id}},
    )
    await db.leagues.update_one(
        {"id": league_id},
        {"$addToSet": {"roster_shooter_ids": shooter_id}},
    )

    return {
        "success": True,
        "message": f"Added '{shooter.get('name')}' to league '{league.get('name')}'",
        "league_id": league_id,
        "shooter_id": shooter_id,
    }


@api_router.get("/matches/{match_id}/roster", response_model=MatchRosterResponse)
async def get_match_roster(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    """Return formal roster plus any shooters who have scores but aren't rostered."""
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    roster_ids = list(match.get("roster_shooter_ids") or [])

    # Score counts for this match by shooter
    pipeline = [
        {"$match": {"match_id": match_id}},
        {"$group": {"_id": "$shooter_id", "count": {"$sum": 1}}},
    ]
    score_counts: Dict[str, int] = {}
    async for row in db.scores.aggregate(pipeline):
        score_counts[row["_id"]] = row["count"]

    async def load_member(sid: str) -> Optional[MatchRosterMember]:
        doc = await db.shooters.find_one({"id": sid})
        if not doc:
            return None
        count = score_counts.get(sid, 0)
        return MatchRosterMember(
            shooter=Shooter(**doc),
            score_count=count,
            has_scores=count > 0,
        )

    members: List[MatchRosterMember] = []
    for sid in roster_ids:
        member = await load_member(sid)
        if member:
            members.append(member)
    members.sort(key=lambda m: (m.shooter.name or "").casefold())

    scored_but_not: List[MatchRosterMember] = []
    roster_set = set(roster_ids)
    for sid, count in score_counts.items():
        if sid not in roster_set:
            member = await load_member(sid)
            if member:
                scored_but_not.append(member)
    scored_but_not.sort(key=lambda m: (m.shooter.name or "").casefold())

    return MatchRosterResponse(
        match_id=match_id,
        members=members,
        scored_but_not_on_roster=scored_but_not,
    )


@api_router.post("/matches/{match_id}/roster", response_model=MatchRosterResponse)
async def add_to_match_roster(
    match_id: str,
    body: MatchRosterAddRequest,
    current_user: User = Depends(get_admin_user),
):
    """
    Admin-only: add existing shooters and/or create new shooters onto this match's roster.
    Does not create scores. Does not remove anyone.
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    if not body.shooter_ids and not body.new_shooters:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail="Provide shooter_ids and/or new_shooters",
        )

    ids_to_add: List[str] = []

    for sid in body.shooter_ids:
        if not sid:
            continue
        exists = await db.shooters.find_one({"id": sid})
        if not exists:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail=f"Shooter id not found: {sid}",
            )
        ids_to_add.append(sid)

    for ns in body.new_shooters:
        data = ns.dict()
        name = (data.get("name") or "").strip()
        if not name:
            raise HTTPException(
                status_code=status.HTTP_400_BAD_REQUEST,
                detail="Each new shooter requires a name",
            )
        shooter_obj, _ = await _create_shooter_record(
            name=name,
            nra_number=_empty_to_none(data.get("nra_number")),
            cmp_number=_empty_to_none(data.get("cmp_number")),
            rating=data.get("rating") or None,
            skip_if_duplicate=False,
        )
        ids_to_add.append(shooter_obj.id)

    if ids_to_add:
        await db.matches.update_one(
            {"id": match_id},
            {"$addToSet": {"roster_shooter_ids": {"$each": ids_to_add}}},
        )

    return await get_match_roster(match_id, current_user)


@api_router.delete("/matches/{match_id}/roster/{shooter_id}")
async def remove_from_match_roster(
    match_id: str,
    shooter_id: str,
    delete_scores: bool = False,
    current_user: User = Depends(get_admin_user),
):
    """
    Admin-only: remove a shooter from this match's roster.

    By default only removes from roster (global shooter profile stays).
    If delete_scores=true, also deletes that shooter's scores for THIS match only.
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    score_count = await db.scores.count_documents(
        {"match_id": match_id, "shooter_id": shooter_id}
    )
    if score_count > 0 and not delete_scores:
        raise HTTPException(
            status_code=status.HTTP_400_BAD_REQUEST,
            detail=(
                f"Shooter has {score_count} score(s) in this match. "
                "Pass delete_scores=true to remove them from the roster and "
                "delete their scores for this match only, or leave them on the roster."
            ),
        )

    deleted_scores = 0
    if delete_scores and score_count > 0:
        result = await db.scores.delete_many(
            {"match_id": match_id, "shooter_id": shooter_id}
        )
        deleted_scores = result.deleted_count

    await db.matches.update_one(
        {"id": match_id},
        {"$pull": {"roster_shooter_ids": shooter_id}},
    )

    return {
        "success": True,
        "deleted_scores": deleted_scores,
        "message": "Removed from match roster"
        + (f" and deleted {deleted_scores} score(s)" if deleted_scores else ""),
    }


@api_router.delete("/matches/{match_id}")
async def delete_match(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    # Only admins can delete matches
    if current_user.role != UserRole.ADMIN: # Use UserRole.ADMIN for consistency
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, 
            detail="Not authorized to delete matches"
        )
    
    # Find the match to ensure it exists
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")
    
    # Delete all scores associated with this match
    delete_scores_result = await db.scores.delete_many({"match_id": match_id})
    
    # Delete match configuration
    await db.match_configs.delete_many({"match_id": match_id})
    
    # Delete the match itself
    delete_match_result = await db.matches.delete_one({"id": match_id})
    
    if delete_match_result.deleted_count == 0:
        raise HTTPException(status_code=500, detail="Failed to delete match")
    
    return {
        "success": True,
        "message": f"Match deleted successfully along with {delete_scores_result.deleted_count} related scores"
    }


@api_router.get("/match-types")
async def get_match_types(current_user: User = Depends(get_current_active_user)):
    """Get all available match types with their stage definitions"""
    result = {}
    for match_type in BasicMatchType:
        stages_info = get_stages_for_match_type(match_type)
        result[match_type] = {
            "entry_stages": stages_info["entry_stages"],
            "subtotal_stages": stages_info["subtotal_stages"],
            "subtotal_mappings": stages_info["subtotal_mappings"],
            "max_score": stages_info["max_score"],
        }
    return result


@api_router.get("/match-config/{match_id}")
async def get_match_config(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    """Get the detailed configuration for a match, including all stages for each match type"""
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    match_obj = Match(**match)

    # Build detailed configuration
    config = {
        "match_id": match_obj.id,
        "match_name": match_obj.name,
        "date": match_obj.date,
        "location": match_obj.location,
        "aggregate_type": match_obj.aggregate_type,
        "match_types": [],
    }

    for match_type_instance in match_obj.match_types:
        stages_info = get_stages_for_match_type(match_type_instance.type)
        config["match_types"].append(
            {
                "type": match_type_instance.type,
                "instance_name": match_type_instance.instance_name,
                "calibers": match_type_instance.calibers,
                "entry_stages": stages_info["entry_stages"],
                "subtotal_stages": stages_info["subtotal_stages"],
                "subtotal_mappings": stages_info["subtotal_mappings"],
                "max_score": stages_info["max_score"],
            }
        )

    return config


# Score Routes
@api_router.post("/scores", response_model=Score)
async def create_score(
    score: ScoreCreate, current_user: User = Depends(get_current_active_user)
):
    # Only admins can create scores
    if current_user.role == UserRole.REPORTER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions"
        )

    # Get match configuration to determine if we need to add subtotals
    match = await db.matches.find_one({"id": score.match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    match_obj = Match(**match)
    match_type_instance = next(
        (
            mt
            for mt in match_obj.match_types
            if mt.instance_name == score.match_type_instance
        ),
        None,
    )

    if not match_type_instance:
        raise HTTPException(status_code=400, detail="Invalid match type instance")

    # Get the match type configuration
    stages_info = get_stages_for_match_type(match_type_instance.type)

    # Calculate total score and X count from the entry stages, skipping null values
    has_valid_score = any(stage.score is not None for stage in score.stages)
    has_valid_x = any(stage.x_count is not None for stage in score.stages)
    
    # If all stages are NULL, mark as not shot and set totals to NULL
    not_shot = not has_valid_score
    total_score = sum(stage.score for stage in score.stages if stage.score is not None) if has_valid_score else None
    total_x_count = sum(stage.x_count for stage in score.stages if stage.x_count is not None) if has_valid_x else None
    
    # Create the score object with calculated totals and not_shot flag
    score_dict = score.dict()
    score_dict.update({
        "total_score": total_score, 
        "total_x_count": total_x_count,
        "not_shot": not_shot
    })

    score_obj = Score(**score_dict)
    await db.scores.insert_one(score_obj.dict())
    return score_obj


@api_router.put("/scores/{score_id}", response_model=Score)
async def update_score(
    score_id: str,
    score_update: ScoreCreate,
    current_user: User = Depends(get_current_active_user),
):
    # Only admins can update scores
    if current_user.role == UserRole.REPORTER:
        raise HTTPException(
            status_code=status.HTTP_403_FORBIDDEN, detail="Not enough permissions"
        )

    # Check if score exists
    existing_score = await db.scores.find_one({"id": score_id})
    if not existing_score:
        raise HTTPException(status_code=404, detail="Score not found")

    # Get match configuration
    match = await db.matches.find_one({"id": score_update.match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")

    match_obj = Match(**match)
    match_type_instance = next(
        (
            mt
            for mt in match_obj.match_types
            if mt.instance_name == score_update.match_type_instance
        ),
        None,
    )

    if not match_type_instance:
        raise HTTPException(status_code=400, detail="Invalid match type instance")

    # Calculate total score and X count from the entry stages, skipping null values
    has_valid_score = any(stage.score is not None for stage in score_update.stages)
    has_valid_x = any(stage.x_count is not None for stage in score_update.stages)
    
    # If all stages are NULL, mark as not shot and set totals to NULL
    not_shot = not has_valid_score
    total_score = sum(stage.score for stage in score_update.stages if stage.score is not None) if has_valid_score else None
    total_x_count = sum(stage.x_count for stage in score_update.stages if stage.x_count is not None) if has_valid_x else None
    
    # Update the score object with calculated totals and not_shot flag
    score_dict = score_update.dict()
    score_dict.update({
        "total_score": total_score, 
        "total_x_count": total_x_count,
        "not_shot": not_shot
    })

    # Update score
    await db.scores.update_one({"id": score_id}, {"$set": score_dict})

    # Get updated score
    updated_score = await db.scores.find_one({"id": score_id})
    return Score(**updated_score)


@api_router.get("/scores", response_model=List[Score])
async def get_scores(
    match_id: Optional[str] = None,
    shooter_id: Optional[str] = None,
    current_user: User = Depends(get_current_active_user),
):
    # Build query
    query = {}
    if match_id:
        query["match_id"] = match_id
    if shooter_id:
        query["shooter_id"] = shooter_id

    scores = await db.scores.find(query).to_list(1000)
    return [Score(**score) for score in scores]


@api_router.get("/scores/{score_id}", response_model=Score)
async def get_score(
    score_id: str, current_user: User = Depends(get_current_active_user)
):
    score = await db.scores.find_one({"id": score_id})
    if not score:
        raise HTTPException(status_code=404, detail="Score not found")
    return Score(**score)


# Special Reports
@api_router.get("/match-report/{match_id}", response_model=Dict[str, Any])
async def get_match_report(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    # Get match details
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")
    
    match_obj = Match(**match)
    
    # Get all scores for this match
    scores = await db.scores.find({"match_id": match_id}).to_list(1000)
    
    # Build shooter data map
    shooter_ids = set(score["shooter_id"] for score in scores)
    shooters = {}
    for shooter_id in shooter_ids:
        shooter = await db.shooters.find_one({"id": shooter_id})
        if shooter:
            shooters[shooter_id] = Shooter(**shooter)
    
    # Get match configuration for subtotal calculations
    match_config = await get_match_config(match_id, current_user)
    match_types_config = {mt["instance_name"]: mt for mt in match_config["match_types"]}
    
    # Organize scores by shooter, match type instance, and caliber
    result = {
        "match": match_obj,
        "shooters": {}
    }
    
    for score in scores:
        score_obj = Score(**score)
        shooter_id = score_obj.shooter_id
        if shooter_id not in result["shooters"] and shooter_id in shooters:
            result["shooters"][shooter_id] = {
                "shooter": shooters[shooter_id],
                "scores": {}
            }
        
        if shooter_id in result["shooters"]:
            shooter_data = result["shooters"][shooter_id]
            match_type_instance = score_obj.match_type_instance
            caliber = score_obj.caliber
            
            key = f"{match_type_instance}_{caliber}"
            
            # Add calculated subtotals if this match type has them
            if match_type_instance in match_types_config:
                mt_config = match_types_config[match_type_instance]
                match_type = mt_config["type"]
                stages_config = get_stages_for_match_type(match_type)
                
                # Use the core function to calculate subtotals
                subtotals = calculate_score_subtotals(score_obj, stages_config)
                
                shooter_data["scores"][key] = {
                    "score": {
                        "id": score_obj.id,
                        "match_type_instance": match_type_instance,
                        "caliber": caliber,
                        "total_score": score_obj.total_score,
                        "total_x_count": score_obj.total_x_count,
                        "not_shot": score_obj.not_shot,  # Include not_shot flag
                        "stages": [
                            {
                                "name": stage.name,
                                "score": stage.score,
                                "x_count": stage.x_count
                            }
                            for stage in score_obj.stages
                        ]
                    },
                    "subtotals": subtotals
                }
    
    # Calculate and add aggregates if applicable
    if match_obj.aggregate_type != "None":
        for shooter_id, shooter_data in result["shooters"].items():
            shooter_scores = shooter_data["scores"]
            shooter_data["aggregates"] = calculate_aggregates(shooter_scores, match_obj)

            # --- Add fallback aggregate total if not present ---
            agg_label = None
            agg_count = 0
            if match_obj.aggregate_type == AggregateType.TWENTY_SEVEN_HUNDRED:
                agg_label = "2700"
                agg_count = 3
            elif match_obj.aggregate_type == AggregateType.EIGHTEEN_HUNDRED_2X900:
                agg_label = "1800"
                agg_count = 2
            elif match_obj.aggregate_type == AggregateType.EIGHTEEN_HUNDRED_3X600:
                agg_label = "1800"
                agg_count = 3

            if agg_label:
                # Find all 900 or 600 scores for this shooter
                scores_list = []
                x_counts = []
                for key, score_data in shooter_scores.items():
                    score = score_data["score"]
                    if score["total_score"] is not None:
                        scores_list.append(score["total_score"])
                        x_counts.append(score["total_x_count"] or 0)
                if len(scores_list) >= agg_count:
                    total = sum(sorted(scores_list, reverse=True)[:agg_count])
                    x_total = sum(sorted(x_counts, reverse=True)[:agg_count])
                    shooter_data["aggregates"][agg_label] = {
                        "score": total,
                        "x_count": x_total,
                    }

    # Include match configuration in the result
    result["match_config"] = match_config
    
    return result


def _shooter_cats(shooter: Shooter) -> List[str]:
    cats = getattr(shooter, "special_categories", None) or []
    out = []
    for c in cats:
        out.append(c.value if hasattr(c, "value") else str(c))
    return out


def _shooter_division(shooter: Shooter) -> Optional[str]:
    d = getattr(shooter, "division", None)
    if d is None:
        return "Civilian"
    return d.value if hasattr(d, "value") else str(d)


def _shooter_rating(shooter: Shooter) -> Optional[str]:
    r = getattr(shooter, "rating", None)
    if r is None:
        return None
    return r.value if hasattr(r, "value") else str(r)


async def _build_bulletin_results_for_event(
    match_id: str,
    *,
    event_scope: str,
    caliber: Optional[str] = None,
    match_type_instance: Optional[str] = None,
) -> List[CompetitorResult]:
    """
    event_scope:
      slow | timed | rapid | total  — filter to instance+caliber scorecard
      caliber_aggregate — sum all totals for caliber
      grand_aggregate — sum all totals for shooter in match
    """
    scores = await db.scores.find({"match_id": match_id}).to_list(5000)
    shooter_ids = {s["shooter_id"] for s in scores}
    shooters: Dict[str, Shooter] = {}
    for sid in shooter_ids:
        doc = await db.shooters.find_one({"id": sid})
        if doc:
            # Tolerate older docs missing new fields
            doc.setdefault("division", "Civilian")
            doc.setdefault("special_categories", [])
            doc.setdefault("competitor_number", None)
            shooters[sid] = Shooter(**doc)

    results: List[CompetitorResult] = []

    if event_scope in ("slow", "timed", "rapid", "nmc", "total"):
        if not caliber or not match_type_instance:
            raise HTTPException(
                status_code=400,
                detail="caliber and match_type_instance are required for this event_scope",
            )
        for s in scores:
            if s.get("match_type_instance") != match_type_instance:
                continue
            cal = s.get("caliber")
            cal_val = cal.value if hasattr(cal, "value") else str(cal)
            if cal_val != caliber:
                continue
            sc, xc = event_score_from_score_doc(s, event_scope)
            if sc is None:
                continue
            sh = shooters.get(s["shooter_id"])
            if not sh:
                continue
            results.append(
                CompetitorResult(
                    shooter_id=sh.id,
                    name=sh.name,
                    competitor_number=getattr(sh, "competitor_number", None),
                    rating=_shooter_rating(sh),
                    division=_shooter_division(sh),
                    special_categories=_shooter_cats(sh),
                    score=sc,
                    x_count=xc or 0,
                )
            )
        return results

    # Aggregates: accumulate per shooter
    totals: Dict[str, Dict[str, int]] = {}
    for s in scores:
        if s.get("not_shot") or s.get("total_score") is None:
            continue
        cal = s.get("caliber")
        cal_val = cal.value if hasattr(cal, "value") else str(cal)
        if event_scope == "caliber_aggregate":
            if not caliber or cal_val != caliber:
                continue
        elif event_scope != "grand_aggregate":
            raise HTTPException(status_code=400, detail=f"Unknown event_scope: {event_scope}")

        sid = s["shooter_id"]
        if sid not in totals:
            totals[sid] = {"score": 0, "x": 0}
        totals[sid]["score"] += int(s["total_score"])
        totals[sid]["x"] += int(s.get("total_x_count") or 0)

    for sid, t in totals.items():
        sh = shooters.get(sid)
        if not sh:
            continue
        results.append(
            CompetitorResult(
                shooter_id=sh.id,
                name=sh.name,
                competitor_number=getattr(sh, "competitor_number", None),
                rating=_shooter_rating(sh),
                division=_shooter_division(sh),
                special_categories=_shooter_cats(sh),
                score=t["score"],
                x_count=t["x"],
            )
        )
    return results


def _event_title(event_scope: str, caliber: Optional[str], match_type_instance: Optional[str], mt_type: Optional[str]) -> str:
    cal = caliber or ""
    if event_scope == "grand_aggregate":
        return "GRAND AGGREGATE MATCH"
    if event_scope == "caliber_aggregate":
        return f"{cal} AGGREGATE MATCH"
    if event_scope == "slow":
        return f"{cal} SLOW FIRE MATCH"
    if event_scope == "timed":
        return f"{cal} TIMED FIRE MATCH"
    if event_scope == "rapid":
        return f"{cal} RAPID FIRE MATCH"
    if event_scope == "nmc":
        return f"{cal} NMC MATCH"
    # full scorecard total (separate from NMC-named event when type is NMC)
    if mt_type == "NMC" or (match_type_instance and "NMC" in (match_type_instance or "").upper()):
        return f"{cal} NMC MATCH"
    if mt_type == "Presidents":
        return f"{cal} PRESIDENTS MATCH"
    if mt_type in ("900", "600"):
        return f"{cal} {mt_type} AGGREGATE"
    return f"{cal} MATCH"


@api_router.get("/match-report/{match_id}/bulletin/events")
async def list_bulletin_events(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    """List available NRA-style bulletin events for this match (for the UI picker)."""
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")
    match_obj = Match(**match)

    events: List[Dict[str, Any]] = []
    n = 1
    for mt in match_obj.match_types:
        mt_type = mt.type.value if hasattr(mt.type, "value") else str(mt.type)
        for cal in mt.calibers:
            cal_v = cal.value if hasattr(cal, "value") else str(cal)
            base = {
                "match_type_instance": mt.instance_name,
                "caliber": cal_v,
                "type": mt_type,
            }
            # SF / TF / RF are separate from NMC (do not include SFNMC/TFNMC/RFNMC in SF/TF/RF)
            for scope, label in (
                ("slow", "Slow Fire"),
                ("timed", "Timed Fire"),
                ("rapid", "Rapid Fire"),
            ):
                events.append(
                    {
                        **base,
                        "event_scope": scope,
                        "match_no": n,
                        "label": f"{cal_v} {label} ({mt.instance_name})",
                        "event_title": _event_title(scope, cal_v, mt.instance_name, mt_type),
                    }
                )
                n += 1
            # NMC as its own event:
            # - NMC course type → full card total
            # - 900 course → mid-block SFNMC+TFNMC+RFNMC only
            events.append(
                {
                    **base,
                    "event_scope": "nmc" if mt_type == "900" else "total",
                    "match_no": n,
                    "label": f"{cal_v} NMC ({mt.instance_name})",
                    "event_title": _event_title("nmc", cal_v, mt.instance_name, mt_type),
                }
            )
            n += 1
            # Full instance total when not already covered as pure NMC course
            if mt_type not in ("NMC",):
                events.append(
                    {
                        **base,
                        "event_scope": "total",
                        "match_no": n,
                        "label": f"{cal_v} {mt.instance_name} Full Total",
                        "event_title": _event_title("total", cal_v, mt.instance_name, mt_type),
                    }
                )
                n += 1

    # Caliber aggregates
    calibers = sorted(
        {
            (c.value if hasattr(c, "value") else str(c))
            for mt in match_obj.match_types
            for c in mt.calibers
        }
    )
    for cal_v in calibers:
        events.append(
            {
                "event_scope": "caliber_aggregate",
                "caliber": cal_v,
                "match_type_instance": None,
                "match_no": n,
                "label": f"{cal_v} Aggregate (all events)",
                "event_title": _event_title("caliber_aggregate", cal_v, None, None),
            }
        )
        n += 1

    events.append(
        {
            "event_scope": "grand_aggregate",
            "caliber": None,
            "match_type_instance": None,
            "match_no": n,
            "label": "Grand Aggregate",
            "event_title": "GRAND AGGREGATE MATCH",
        }
    )
    return {"match_id": match_id, "events": events}


@api_router.get("/match-report/{match_id}/bulletin")
async def get_match_bulletin(
    match_id: str,
    event_scope: str = "total",
    caliber: Optional[str] = None,
    match_type_instance: Optional[str] = None,
    match_no: int = 1,
    current_user: User = Depends(get_current_active_user),
):
    """
    NRA Tournament Results Bulletin for one event.

    Query params mirror docs/NRA_BULLETIN_SPEC.md event scopes.
    """
    match = await db.matches.find_one({"id": match_id})
    if not match:
        raise HTTPException(status_code=404, detail="Match not found")
    match_obj = Match(**match)

    mt_type = None
    if match_type_instance:
        for mt in match_obj.match_types:
            if mt.instance_name == match_type_instance:
                mt_type = mt.type.value if hasattr(mt.type, "value") else str(mt.type)
                break

    results = await _build_bulletin_results_for_event(
        match_id,
        event_scope=event_scope,
        caliber=caliber,
        match_type_instance=match_type_instance,
    )

    date_line = match_obj.date.strftime("%B %d, %Y") if match_obj.date else ""
    tournament_title = (
        match_obj.tournament_name
        or (
            f"{'NRA REGISTERED MATCH' if match_obj.is_nra_registered else match_obj.name} -- {date_line}"
        )
    )
    event_title = _event_title(event_scope, caliber, match_type_instance, mt_type)

    bulletin = build_bulletin(
        tournament_title=tournament_title,
        date_line=date_line,
        location=match_obj.location or "",
        match_no=match_no,
        event_title=event_title,
        results=results,
    )
    bulletin["match_id"] = match_id
    bulletin["query"] = {
        "event_scope": event_scope,
        "caliber": caliber,
        "match_type_instance": match_type_instance,
        "match_no": match_no,
    }
    return bulletin


@api_router.get("/match-report/{match_id}/bulletin/excel")
async def get_match_bulletin_excel(
    match_id: str,
    event_scope: str = "total",
    caliber: Optional[str] = None,
    match_type_instance: Optional[str] = None,
    match_no: int = 1,
    current_user: User = Depends(get_current_active_user),
):
    """Excel export of the NRA bulletin (same sections as the web view)."""
    from .excel_style import (
        apply_print_setup,
        autosize_columns,
        font_body,
        font_subtitle,
        font_title,
        style_data_row,
        style_header_row,
        style_section_banner,
        THIN,
        align_center,
    )

    bulletin = await get_match_bulletin(
        match_id,
        event_scope=event_scope,
        caliber=caliber,
        match_type_instance=match_type_instance,
        match_no=match_no,
        current_user=current_user,
    )
    wb = Workbook()
    ws = wb.active
    ws.title = "Bulletin"
    h = bulletin["header"]
    max_col = 5

    def write_cells(values: List[Any]) -> int:
        ws.append(list(values))
        return ws.max_row

    # Title block
    r = write_cells([h["bulletin_title"]])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    c = ws.cell(row=r, column=1)
    c.font = font_title()
    c.alignment = align_center()

    r = write_cells([h["tournament_title"]])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    ws.cell(row=r, column=1).font = font_subtitle()
    ws.cell(row=r, column=1).alignment = align_center()

    r = write_cells([h.get("location") or ""])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    ws.cell(row=r, column=1).alignment = align_center()

    write_cells([])
    r = write_cells([f"MATCH NO. {h['match_no']} -- {h['event_title']}"])
    ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=max_col)
    ws.cell(row=r, column=1).font = font_subtitle()
    ws.cell(row=r, column=1).alignment = align_center()
    write_cells([])

    def write_table_header(cols: List[str]) -> int:
        row = write_cells(cols)
        style_header_row(ws, row, len(cols))
        return row

    def write_place_rows(rows: List[Dict[str, Any]], *, with_place: bool) -> None:
        for i, item in enumerate(rows):
            if with_place:
                values = [
                    item.get("place"),
                    item.get("competitor_number"),
                    item.get("name_display"),
                    item.get("score_display"),
                    item.get("award_label") or "",
                ]
            else:
                values = [
                    "",
                    item.get("competitor_number"),
                    item.get("name_display"),
                    item.get("score_display"),
                    item.get("award_label") or "",
                ]
            row = write_cells(values)
            # Top-3 place awards and labeled class places get gold highlight
            award = item.get("award_label") or ""
            highlight = bool(award) and (
                "Winner" in award
                or "First" in award
                or "Second" in award
                or "Third" in award
                or "Fourth" in award
                or award.startswith("High ")
            )
            special = award.startswith("High ")
            style_data_row(
                ws,
                row,
                max_col,
                alt=(i % 2 == 1) and not highlight and not special,
                highlight=highlight and not special,
                special=special,
            )

    # OPEN place awards
    r = write_cells(
        [f"OPEN -- PLACE AWARDS ({bulletin['competitor_count']} COMPETITORS)"]
    )
    style_section_banner(ws, r, max_col)
    write_table_header(["Place", "Comp #", "Name", "Score", "Award"])
    write_place_rows(bulletin.get("open_place_awards") or [], with_place=True)

    write_cells([])
    r = write_cells(["SPECIAL CATEGORY AWARDS"])
    style_section_banner(ws, r, max_col)
    write_table_header(["", "Comp #", "Name", "Score", "Award"])
    write_place_rows(bulletin.get("special_category_awards") or [], with_place=False)

    for sec in bulletin.get("class_sections") or []:
        write_cells([])
        r = write_cells(
            [f"{sec['title']} ({sec['competitor_count']} COMPETITORS)"]
        )
        style_section_banner(ws, r, max_col)
        write_table_header(["Place", "Comp #", "Name", "Score", "Award"])
        write_place_rows(sec.get("rows") or [], with_place=True)

    # Column widths tuned for bulletin
    widths = [8, 10, 32, 14, 36]
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w

    apply_print_setup(ws, landscape=False, fit_width=True)
    ws.freeze_panes = "A8"

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    safe_event = re.sub(r"[^\w\-]+", "_", h["event_title"])[:40]
    filename = f"bulletin_{safe_event}.xlsx"
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )


@api_router.get("/match-report/{match_id}/excel")
async def get_match_report_excel(
    match_id: str, current_user: User = Depends(get_current_active_user)
):
    # Get the match report data first (reuse existing function)
    report_data = await get_match_report(match_id, current_user)
    match_obj: Match = report_data["match"] # Added type hint
    shooters_data = report_data["shooters"]
    
    # Create a new workbook
    from .excel_style import (
        apply_print_setup,
        fill_alt,
        fill_header,
        font_body,
        font_header,
        font_meta_label,
        font_title,
        THIN,
        align_center,
        align_left,
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Match Report"
    
    # Define styles (aligned with excel_style palette)
    header_font = font_header()
    header_fill = fill_header()
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    thin_border = THIN
    
    # Add match details
    ws.append(["Match Report"])
    ws.merge_cells(f"A1:G1") # Adjust merge range if needed, G1 seems fine for now
    cell = ws.cell(row=1, column=1)
    cell.font = font_title()
    cell.alignment = Alignment(horizontal="center")
    
    ws.append([])
    ws.append(["Match Name:", match_obj.name])
    ws.append(["Date:", match_obj.date.strftime("%Y-%m-%d")])
    ws.append(["Location:", match_obj.location])
    
    agg_type_map = {
        AggregateType.TWENTY_SEVEN_HUNDRED: "2700 (3x900)", # Clarified display
        AggregateType.EIGHTEEN_HUNDRED_2X900: "1800 (2x900)",
        AggregateType.EIGHTEEN_HUNDRED_3X600: "1800 (3x600)",
        AggregateType.NONE: "None"
    }
    agg_type_display = agg_type_map.get(match_obj.aggregate_type, str(match_obj.aggregate_type.value if isinstance(match_obj.aggregate_type, Enum) else match_obj.aggregate_type))
    
    ws.append(["Aggregate Type:", agg_type_display])
    # Bold meta labels (column A)
    for meta_row in (3, 4, 5, 6):
        lab = ws.cell(row=meta_row, column=1)
        lab.font = font_meta_label()
        val = ws.cell(row=meta_row, column=2)
        val.font = font_body(bold=True)
    ws.append([]) # Blank row
    
    is_aggregate = match_obj.aggregate_type != AggregateType.NONE
    
    # Variables for dynamic header content
    header_row1_content = []
    header_row2_content = []
    ordered_calibers_for_agg = []
    agg_sub_fields_for_agg = []
    base_match_type_for_agg_val = None # Store the BasicMatchType enum value

    # --- Build summary header ---
    if is_aggregate:
        header_row1_content, header_row2_content, ordered_calibers_for_agg, agg_sub_fields_for_agg, base_match_type_for_agg_val = \
            _build_dynamic_aggregate_header_and_calibers(match_obj)
        
        # Only append if headers are generated (i.e., valid aggregate type)
        if header_row1_content and header_row2_content:
            ws.append(header_row1_content)
            ws.append(header_row2_content)
        header_row_for_styling_and_cols = header_row2_content # Used for column counts and styling main header
        header_offset = 2 if header_row1_content and header_row2_content else 0
    else:
        header_row_for_styling_and_cols = _build_dynamic_non_aggregate_header(match_obj)
        ws.append(header_row_for_styling_and_cols)
        header_offset = 1

    current_header_start_row = 8 # Assuming match details take up to row 7

    # Add shooter rows
    for idx, (shooter_id, s_data) in enumerate(shooters_data.items()): # s_data to avoid conflict
        shooter_obj = s_data["shooter"]
        row_content_list: List[Any] # Type hint for clarity
        if is_aggregate:
            if base_match_type_for_agg_val and ordered_calibers_for_agg and agg_sub_fields_for_agg: # Ensure all parts are valid
                row_content_list = build_aggregate_row_grouped(
                    shooter_obj, s_data, report_data, 
                    ordered_calibers_for_agg, agg_sub_fields_for_agg, base_match_type_for_agg_val
                )
            else: # Should not happen if headers were generated
                pass
        else:
            row_content_list = build_non_aggregate_row(shooter_obj, s_data, match_obj) # Pass match_obj
        
        for col_idx_excel, value in enumerate(row_content_list, 1):
            ws.cell(row=current_header_start_row + header_offset + idx, column=col_idx_excel, value=value)

    # Apply header styles (for both header rows if aggregate)
    if header_offset > 0: # Check if any header was actually added
        if is_aggregate and header_offset == 2:
            # Determine start columns for caliber names dynamically
            caliber_start_cols_excel = [3] # First caliber starts at column C (3)
            if ordered_calibers_for_agg and agg_sub_fields_for_agg:
                current_col_for_calc = 3 + len(agg_sub_fields_for_agg)
                for _ in range(1, len(ordered_calibers_for_agg)):
                    caliber_start_cols_excel.append(current_col_for_calc)
                    current_col_for_calc += len(agg_sub_fields_for_agg)

            for r_idx_offset in range(header_offset):
                actual_row_idx = current_header_start_row + r_idx_offset
                # Use length of header_row1_content for first header row, header_row2_content for second
                current_row_content = header_row1_content if r_idx_offset == 0 else header_row2_content
                for col_idx_excel in range(1, len(current_row_content) + 1):
                    cell = ws.cell(row=actual_row_idx, column=col_idx_excel)
                    if r_idx_offset == 0: # Caliber row (first header row)
                        if col_idx_excel in caliber_start_cols_excel:
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="left") # Caliber names left-aligned
                            cell.border = thin_border
                        elif col_idx_excel <= 2 : # Shooter, Agg Total in first header row - no special style
                            cell.border = Border() # No border for these in the caliber row
                        else: # Blank cells under merged caliber name
                            cell.border = thin_border # Keep border for structure
                    else: # Fields row (second header row or the only header row for non-agg)
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = header_alignment
                        cell.border = thin_border
        elif not is_aggregate and header_offset == 1: # Non-aggregate single header row
             for col_idx_excel in range(1, len(header_row_for_styling_and_cols) + 1):
                cell = ws.cell(row=current_header_start_row, column=col_idx_excel)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border

    # Merge cells for caliber groupings in the first header row (aggregate only)
    if is_aggregate and header_offset == 2 and ordered_calibers_for_agg and agg_sub_fields_for_agg:
        start_col_merge = 3 # Column C
        for i in range(len(ordered_calibers_for_agg)):
            if len(agg_sub_fields_for_agg) > 1: # Only merge if there's more than one sub-field
                ws.merge_cells(
                    start_row=current_header_start_row, start_column=start_col_merge,
                    end_row=current_header_start_row, end_column=start_col_merge + len(agg_sub_fields_for_agg) - 1
                )
            start_col_merge += len(agg_sub_fields_for_agg)

    # Auto-adjust column widths for headers (use the longest header row for column count)
    # This needs to be based on the actual content of header_row_for_styling_and_cols
    if header_row_for_styling_and_cols:
        ws.column_dimensions[get_column_letter(1)].width = 25 # Shooter name
        ws.column_dimensions[get_column_letter(2)].width = 18 # Aggregate Total / Average
        for i in range(3, len(header_row_for_styling_and_cols) + 1):
             ws.column_dimensions[get_column_letter(i)].width = 12 # Default for score columns

    # Determine columns to bold (e.g., "900" or "600" total columns for aggregates)
    total_col_indices_to_bold = []
    if is_aggregate and agg_sub_fields_for_agg:
        total_field_name = agg_sub_fields_for_agg[-1] # e.g., "900" or "600"
        # header_row_for_styling_and_cols is header_row2_content here
        for i, h_val in enumerate(header_row_for_styling_and_cols):
            if h_val == total_field_name:
                total_col_indices_to_bold.append(i + 1) # 1-indexed

    data_start_excel_row = current_header_start_row + header_offset
    for idx, (shooter_id, s_data) in enumerate(shooters_data.items()): # s_data to avoid conflict
        shooter_obj = s_data["shooter"]
        row_content_list: List[Any] # Type hint for clarity
        if is_aggregate:
            if base_match_type_for_agg_val and ordered_calibers_for_agg and agg_sub_fields_for_agg: # Ensure all parts are valid
                row_content_list = build_aggregate_row_grouped(
                    shooter_obj, s_data, report_data, 
                    ordered_calibers_for_agg, agg_sub_fields_for_agg, base_match_type_for_agg_val
                )
            else: # Should not happen if headers were generated
                pass
        else:
            row_content_list = build_non_aggregate_row(shooter_obj, s_data, match_obj) # Pass match_obj
        
        for col_idx_excel, value in enumerate(row_content_list, 1):
            ws.cell(row=data_start_excel_row + idx, column=col_idx_excel, value=value)

    # Apply borders, zebra striping, and alignment to data cells
    for row_offset, row in enumerate(
        ws.iter_rows(
            min_row=data_start_excel_row,
            max_row=ws.max_row,
            min_col=1,
            max_col=ws.max_column,
        )
    ):
        alt = row_offset % 2 == 1
        for cell in row:
            cell.border = thin_border
            if cell.col_idx > 1:  # Center-align score columns
                cell.alignment = Alignment(horizontal="center")
            else:
                cell.alignment = align_left()
            if alt:
                cell.fill = fill_alt()
            if cell.col_idx in total_col_indices_to_bold:
                cell.font = Font(bold=True)

    # Freeze panes for both aggregate and non-aggregate matches
    if is_aggregate:
        # For aggregate matches, freeze after "Aggregate Total" column (column B)
        ws.freeze_panes = f"C{data_start_excel_row}"
    else:
        # For non-aggregate matches, freeze after "Average" column (column B)
        ws.freeze_panes = f"C{data_start_excel_row}"

    apply_print_setup(ws, landscape=True, fit_width=True)

    # Create detailed sheets for each shooter
    for shooter_id, shooter_data in shooters_data.items():
        shooter = shooter_data["shooter"]
        ws_detail = wb.create_sheet(title=f"{shooter.name[:28]}")  # Limit sheet name length
        
        # Add shooter details
        ws_detail.append(["Shooter Report"])
        ws_detail.merge_cells(f"A1:C1")
        cell = ws_detail.cell(row=1, column=1)
        cell.font = font_title()
        cell.alignment = Alignment(horizontal="center")
        
        ws_detail.append([])
        ws_detail.append(["Shooter Name:", shooter.name])
        ws_detail.append(["Match Name:", match_obj.name])
        ws_detail.append(["Date:", match_obj.date.strftime("%Y-%m-%d")])
        ws_detail.append(["Location:", match_obj.location])
        ws_detail.append(["NRA Number:", shooter.nra_number or "-"])
        ws_detail.append(["CMP Number:", shooter.cmp_number or "-"])
        for meta_row in range(3, 9):
            ws_detail.cell(row=meta_row, column=1).font = font_meta_label()
            ws_detail.cell(row=meta_row, column=2).font = font_body()
        ws_detail.append([])
        apply_print_setup(ws_detail, landscape=False, fit_width=True)
        
        if is_aggregate:
            total_possible_display_value = ""
            agg_main_label_for_lookup = ""
            
            if match_obj.aggregate_type == AggregateType.TWENTY_SEVEN_HUNDRED:
                total_possible_display_value = "2700"
                agg_main_label_for_lookup = "2700"
            elif match_obj.aggregate_type == AggregateType.EIGHTEEN_HUNDRED_2X900:
                total_possible_display_value = "1800"
                agg_main_label_for_lookup = "1800"
            elif match_obj.aggregate_type == AggregateType.EIGHTEEN_HUNDRED_3X600:
                total_possible_display_value = "1800"
                agg_main_label_for_lookup = "1800"

            # Display the main aggregate total for the shooter
            # Try to get this from the pre-calculated aggregates in shooter_data
            main_agg_score_display = "-"
            if agg_main_label_for_lookup and \
               shooter_data.get("aggregates") and \
               shooter_data["aggregates"].get(agg_main_label_for_lookup):
                agg_info = shooter_data["aggregates"][agg_main_label_for_lookup]
                main_agg_score_val = agg_info.get("score", 0)
                main_agg_x_val = agg_info.get("x_count", 0)
                if main_agg_score_val > 0 or main_agg_x_val > 0:
                    main_agg_score_display = f"{main_agg_score_val} ({main_agg_x_val}X)"
            
            ws_detail.append([f"Aggregate Total ({total_possible_display_value}):", main_agg_score_display])

            # Per-caliber breakdown for the aggregate's components
            base_match_type_for_agg_detail, agg_sub_fields_detail = _get_aggregate_components(match_obj.aggregate_type)
            ordered_calibers_for_agg_detail = []
            if base_match_type_for_agg_detail:
                 ordered_calibers_for_agg_detail = _get_ordered_calibers_for_aggregate(match_obj, base_match_type_for_agg_detail)

            if ordered_calibers_for_agg_detail and base_match_type_for_agg_detail:
                # Determine if the sub-total is per 900 or 600 points
                sub_total_points_label = ""
                if base_match_type_for_agg_detail == BasicMatchType.NINEHUNDRED:
                    sub_total_points_label = "900"
                elif base_match_type_for_agg_detail == BasicMatchType.SIXHUNDRED:
                    sub_total_points_label = "600"
                
                if sub_total_points_label: # Only proceed if we have a valid label (900 or 600)
                    for caliber_enum_detail in ordered_calibers_for_agg_detail:
                        caliber_str_detail = caliber_enum_detail.value
                        caliber_component_total_score = 0
                        caliber_component_total_x = 0
                        has_data_for_caliber_component = False

                        # Sum scores for this caliber that are of the aggregate's base match type
                        for score_key, score_item_detail in shooter_data["scores"].items():
                            score_details_item = score_item_detail["score"]
                            if score_details_item["caliber"] == caliber_str_detail:
                                # Check if this score's match_type_instance is of the base_match_type_for_agg_detail
                                is_part_of_aggregate_base_type = False
                                for mt_cfg in match_obj.match_types:
                                    if mt_cfg.instance_name == score_details_item["match_type_instance"] and \
                                       mt_cfg.type == base_match_type_for_agg_detail:
                                        is_part_of_aggregate_base_type = True
                                        break
                                
                                if is_part_of_aggregate_base_type:
                                    if not score_details_item.get("not_shot", False):
                                        if score_details_item["total_score"] is not None:
                                            caliber_component_total_score += score_details_item["total_score"]
                                            has_data_for_caliber_component = True
                                        if score_details_item["total_x_count"] is not None:
                                            caliber_component_total_x += score_details_item["total_x_count"]
                        
                        display_val = f"{caliber_component_total_score} ({caliber_component_total_x}X)" if has_data_for_caliber_component else "-"
                        ws_detail.append([f"{caliber_str_detail} {sub_total_points_label}", display_val])
            
            ws_detail.append([])  # Blank row before detailed stage breakdown
        
        # For each match type and caliber, add detailed scores
        row_index = 9  # Starting row for score details
        
        for mt in match_obj.match_types:
            match_config = None
            for mt_config in report_data.get("match_config", {}).get("match_types", []):
                if mt_config["instance_name"] == mt.instance_name:
                    match_config = mt_config
                    break
            
            if not match_config:
                continue
                
            stages_config = get_stages_for_match_type(mt.type)
            
            for caliber in mt.calibers:
                # Try multiple key formats as in the summary
                key_formats = [
                    f"{mt.instance_name}_{caliber.value}", # Use caliber.value here
                    f"{mt.instance_name}_CaliberType.{caliber.value.replace('.', '').upper()}" # And here
                ]
                
                # Add special cases by comparing caliber.value
                if caliber == CaliberType.TWENTYTWO: # Or caliber.value == ".22"
                    key_formats.append(f"{mt.instance_name}_CaliberType.TWENTYTWO")
                elif caliber == CaliberType.CENTERFIRE: # Or caliber.value == "CF"
                    key_formats.append(f"{mt.instance_name}_CaliberType.CENTERFIRE")
                elif caliber == CaliberType.FORTYFIVE: # Or caliber.value == ".45"
                    key_formats.append(f"{mt.instance_name}_CaliberType.FORTYFIVE")
                elif caliber == CaliberType.SERVICEPISTOL: # Or caliber.value == "Service Pistol"
                    key_formats.append(f"{mt.instance_name}_CaliberType.SERVICEPISTOL")
                    key_formats.append(f"{mt.instance_name}_CaliberType.NINESERVICE")
                    key_formats.append(f"{mt.instance_name}_CaliberType.FORTYFIVESERVICE")
                elif caliber == CaliberType.SERVICEREVOLVER: # Or caliber.value == "Service Revolver"
                    key_formats.append(f"{mt.instance_name}_CaliberType.SERVICEREVOLVER")
                elif caliber == CaliberType.DR: # Or caliber.value == "DR"
                    key_formats.append(f"{mt.instance_name}_CaliberType.DR")
                
                # Try to find a matching score
                score_data = None
                for key in key_formats:
                    if key in shooter_data["scores"]:
                        score_data = shooter_data["scores"][key]
                        break
                
                if not score_data:
                    continue
                
                # Add header for this match type and caliber
                ws_detail.append([f"{mt.instance_name} - {caliber.value}"]) # Use caliber.value for display
                current_row = ws_detail.max_row  # Get the actual row that was just appended
                
                # Add header for this match type and caliber
                ws_detail.append([f"{mt.instance_name} - {caliber.value}"]) # Use caliber.value for display
                current_row = ws_detail.max_row  # Get the actual row that was just appended
                
                # Apply filled background to header row
                for col in range(1, 4):
                    cell = ws_detail.cell(row=current_row, column=col)
                    cell.fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
                
                # Apply bold font to the first cell which has the header text
                cell = ws_detail.cell(row=current_row, column=1)
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal="center")
                
                row_index += 1
                
                # Add stage headers
                header_row = ["Stage", "Score", "X Count"]
                ws_detail.append(header_row)
                current_row = ws_detail.max_row  # Get the actual row for the header
                
                # Apply header styles
                for col in range(1, len(header_row) + 1):
                    cell = ws_detail.cell(row=current_row, column=col)
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
                    cell.border = thin_border
                
                row_index += 1
                
                # Check if this is a not_shot match
                not_shot = score_data["score"].get("not_shot", False)
                if not_shot:
                    # Add "Not Shot" indicator with special formatting
                    ws_detail.append(["Not Shot"])
                    current_row = ws_detail.max_row
                    not_shot_cell = ws_detail.cell(row=current_row, column=1)
                    not_shot_cell.font = Font(bold=True, color="FF0000")  # Red text
                    ws_detail.merge_cells(f"A{current_row}:C{current_row}")
                    row_index += 1
                    
                    # Add total row with dashes
                    ws_detail.append([
                        "Total",
                        "-",
                        "-"
                    ])
                    current_row = ws_detail.max_row
                    
                    # Apply total row formatting
                    for col in range(1, 4):
                        cell = ws_detail.cell(row=current_row, column=col)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        if col > 1:  # Center-align score columns
                            cell.alignment = Alignment(horizontal="center")
                            
                    row_index += 1
                    
                else:
                    # Add stage scores
                    stages = score_data["score"]["stages"]
                    for stage in stages:
                        stage_name = stage["name"]
                        score_value = stage["score"]
                        x_count = stage["x_count"]
                        
                        # Format the score and x_count correctly for display
                        score_display = "-" if score_value is None else score_value
                        x_count_display = "-" if x_count is None else x_count
                        
                        ws_detail.append([
                            stage_name,
                            score_display,
                            x_count_display
                        ])
                        current_row = ws_detail.max_row
                        
                        # Apply borders to data cells
                        for col in range(1, len(header_row) + 1):
                            cell = ws_detail.cell(row=current_row, column=col)
                            cell.border = thin_border
                            if col > 1:  # Align score columns to center
                                cell.alignment = Alignment(horizontal="center")
                        
                        row_index += 1
                    
                    # Add total row
                    total_score = score_data["score"]["total_score"]
                    total_x_count = score_data["score"]["total_x_count"]
                    
                    # Format the total score and x_count correctly for display
                    total_score_display = "-" if total_score is None else total_score
                    total_x_count_display = "-" if total_x_count is None else total_x_count
                    
                    ws_detail.append([
                        "Total",
                        total_score_display,
                        total_x_count_display
                    ])
                    current_row = ws_detail.max_row
                
                    # Apply total row styling
                    for col in range(1, len(header_row) + 1):
                        cell = ws_detail.cell(row=current_row, column=col)
                        cell.font = Font(bold=True)
                        cell.border = thin_border
                        if col > 1:  # Align score columns to center
                            cell.alignment = Alignment(horizontal="center")
                
                row_index += 2  # Space before next match type
        
        # Auto-adjust column widths
        for i, column_width in enumerate([15, 10, 10], 1):
            ws_detail.column_dimensions[get_column_letter(i)].width = column_width
    
    # Save to BytesIO object
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)  # Rewind to beginning
    
    filename = f"match_report_{match_obj.name.replace(' ', '_')}_{match_obj.date.strftime('%Y-%m-%d')}.xlsx"
    
    headers = {
        "Content-Disposition": f"attachment; filename={filename}",
        "Access-Control-Expose-Headers": "Content-Disposition"  # Important for CORS
    }
    
    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )


# Add shooter averages endpoint for ShooterDetail component
@api_router.get("/shooter-averages/{shooter_id}")
async def get_shooter_averages(
    shooter_id: str, current_user: User = Depends(get_current_active_user)
):
    """Get shooter's average performance by caliber"""
    shooter = await db.shooters.find_one({"id": shooter_id})
    if not shooter:
        raise HTTPException(status_code=404, detail="Shooter not found")

    # Get all scores for this shooter
    scores = await db.scores.find({"shooter_id": shooter_id}).to_list(1000)
    score_objects = [Score(**score) for score in scores]
    
    # Use the core function to calculate averages
    averages = calculate_shooter_averages_by_caliber(score_objects)

    return {"caliber_averages": averages}

@api_router.get("/shooter-report/{shooter_id}")
async def get_shooter_report(
    shooter_id: str, current_user: User = Depends(get_current_active_user)
):
    """Get comprehensive shooter report including matches and detailed scores with averages"""
    
    shooter = await db.shooters.find_one({"id": shooter_id})
    if not shooter:
        raise HTTPException(status_code=404, detail="Shooter not found")

    shooter_obj = Shooter(**shooter)

    # Get all scores for this shooter
    scores = await db.scores.find({"shooter_id": shooter_id}).to_list(1000)
    score_objs = [Score(**score) for score in scores]

    # Get all matches the shooter participated in
    match_ids = set(score.match_id for score in score_objs)
    matches = {}
    for match_id in match_ids:
        match = await db.matches.find_one({"id": match_id})
        if match:
            matches[match_id] = Match(**match)

    # Build detailed report
    report = {
        "shooter": shooter_obj,
        "matches": {},
        "averages": {"by_match_type": {}, "by_caliber": {}},
    }

    # Group scores by match
    for score in score_objs:
        match_id = score.match_id
        if match_id in matches:
            match = matches[match_id]
            if match_id not in report["matches"]:
                report["matches"][match_id] = {"match": match, "scores": []}

            # Add score details
            report["matches"][match_id]["scores"].append(
                {
                    "score": score,
                    "match_type": next(
                        (
                            mt
                            for mt in match.match_types
                            if mt.instance_name == score.match_type_instance
                        ),
                        None,
                    ),
                }
            )

    # Calculate averages by match type and caliber
    averages_by_type = {}
    averages_by_caliber = {}

    for score in score_objs:
        match_id = score.match_id
        if match_id in matches:
            match = matches[match_id]
            match_type = next(
                (
                    mt.type
                    for mt in match.match_types
                    if mt.instance_name == score.match_type_instance
                ),
                None,
            )

            if match_type:
                # By match type and caliber
                key = f"{match_type}_{score.caliber}"
                if key not in averages_by_type:
                    averages_by_type[key] = {
                        "count": 0,
                        "total_score": 0,
                        "total_x_count": 0,
                        "stages": {},
                    }

                # Skip NULL scores and scores marked as not_shot
                if score.total_score is None or getattr(score, "not_shot", False):
                    continue
                    
                avg_data = averages_by_type[key]
                avg_data["count"] += 1
                avg_data["total_score"] += score.total_score
                avg_data["total_x_count"] += (score.total_x_count or 0)

                # Track stage scores
                for stage in score.stages:
                    if stage.name not in avg_data["stages"]:
                        avg_data["stages"][stage.name] = {
                            "score_sum": 0,
                            "x_count_sum": 0,
                        }

                    avg_data["stages"][stage.name]["score_sum"] += (stage.score if stage.score is not None else 0)
                    avg_data["stages"][stage.name]["count"] = avg_data["stages"][stage.name].get("count", 0) + (1 if stage.score is not None else 0)
                    avg_data["stages"][stage.name]["x_count_sum"] += (stage.x_count if stage.x_count is not None else 0)

                # By caliber only
                if score.caliber not in averages_by_caliber:
                    averages_by_caliber[score.caliber] = {
                        "count": 0,
                        "total_score_sum": 0,
                        "total_x_count_sum": 0,
                        "match_types": {},
                    }

                cal_data = averages_by_caliber[score.caliber]
                cal_data["count"] += 1
                cal_data["total_score_sum"] += score.total_score
                cal_data["total_x_count_sum"] += (score.total_x_count or 0)

                # Track match type data
                if match_type not in cal_data["match_types"]:
                    cal_data["match_types"][match_type] = {
                        "count": 0,
                        "score_sum": 0,
                        "x_count_sum": 0,
                    }

                cal_data["match_types"][match_type]["count"] += 1
                cal_data["match_types"][match_type]["score_sum"] += score.total_score
                cal_data["match_types"][match_type]["x_count_sum"] += (score.total_x_count or 0)

    # Calculate final averages
    for key, data in averages_by_type.items():
        if data["count"] > 0:
            valid_scores_count = sum(1 for stage_name, stage_data in data["stages"].items() if stage_data.get("count", 0) > 0)
            if valid_scores_count > 0:
                data["avg_score"] = round(data["total_score"] / data["count"], 2)
                data["avg_x_count"] = round(data["total_x_count"] / data["count"], 2)

                for stage_name, stage_data in data["stages"].items():
                    stage_count = stage_data.get("count", 0)
                    if stage_count > 0:
                        stage_data["avg_score"] = round(
                            stage_data["score_sum"] / stage_count, 2
                        )
                        stage_data["avg_x_count"] = round(
                            stage_data["x_count_sum"] / stage_count, 2
                        )

    for caliber, data in averages_by_caliber.items():
        if data["count"] > 0:
            data["avg_score"] = round(data["total_score_sum"] / data["count"], 2)
            data["avg_x_count"] = round(data["total_x_count_sum"] / data["count"], 2)

            for match_type, mt_data in data["match_types"].items():
                if mt_data["count"] > 0:
                    mt_data["avg_score"] = round(
                        mt_data["score_sum"] / mt_data["count"], 2
                    )
                    mt_data["avg_x_count"] = round(
                        mt_data["x_count_sum"] / mt_data["count"], 2
                    )

    report["averages"]["by_match_type"] = averages_by_type
    report["averages"]["by_caliber"] = averages_by_caliber

    return report

@api_router.get("/")
async def root():
    return {"message": "Enhanced Shooting Match Score Management API"}


# Include the routers in the main app
app.include_router(auth_router, prefix="/api", tags=["Authentication"])
app.include_router(api_router, tags=["API"])

# Get allowed origins from environment variable
origins_env = os.environ.get("ORIGINS", "")

# Use only environment-provided origins or a minimal fallback
if origins_env:
    allowed_origins = [origin.strip() for origin in origins_env.split(",")]
else:
    # Minimal fallback for development when no ORIGINS env var is set
    allowed_origins = ["http://localhost:3000", "http://localhost:8080"]

app.add_middleware(
    CORSMiddleware,
    allow_credentials=True,
    allow_origins=allowed_origins,
    allow_methods=["*"],
    allow_headers=["*"],
    expose_headers=["*"],
)

async def create_first_admin():
    """Seed a default admin when the users collection is empty."""
    try:
        users_count = await db.users.count_documents({})

        if users_count == 0:
            default_email = "admin@example.com"
            default_password = "admin123"  # Change this in production!
            hashed_password = get_password_hash(default_password)

            user = UserInDB(
                id=str(uuid.uuid4()),
                email=default_email,
                username="admin",
                role=UserRole.ADMIN,
                hashed_password=hashed_password,
                created_at=datetime.utcnow(),
                is_active=True,
            )

            user_dict = user.dict()
            await db.users.insert_one(user_dict)
            logger.info(f"Created default admin user: {default_email}")
            logger.info(f"Admin user ID: {user_dict['id']}")
        else:
            logger.info(
                f"Database already has {users_count} users, skipping default admin creation"
            )
    except Exception as e:
        logger.error(f"Error creating default admin user: {str(e)}")
        # Don't fail startup; log and continue so the API still serves


@app.on_event("startup")
async def startup_event():
    await connect_to_mongo()
    await create_first_admin()


@app.on_event("shutdown")
async def shutdown_event():
    await close_mongo_connection()

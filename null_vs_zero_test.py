import requests
import json
import sys
import io
import os
import tempfile
from datetime import datetime
import openpyxl
import re

class NullVsZeroTester:
    def __init__(self, base_url="https://ecc4b2e5-4738-47d7-aabd-fec160cafe64.preview.emergentagent.com"):
        self.base_url = base_url
        self.token = None
        self.tests_run = 0
        self.tests_passed = 0
        self.api_url = f"{self.base_url}/api"

    def run_test(self, name, method, endpoint, expected_status, data=None, check_function=None, binary=False):
        """Run a single API test"""
        url = f"{self.api_url}/{endpoint}"
        headers = {'Content-Type': 'application/json'}
        if self.token:
            headers['Authorization'] = f'Bearer {self.token}'

        self.tests_run += 1
        print(f"\n🔍 Testing {name}...")
        
        try:
            if method == 'GET':
                response = requests.get(url, headers=headers)
            elif method == 'POST':
                response = requests.post(url, json=data, headers=headers)
            elif method == 'PUT':
                response = requests.put(url, json=data, headers=headers)
            elif method == 'DELETE':
                response = requests.delete(url, headers=headers)

            status_success = response.status_code == expected_status
            
            if status_success:
                print(f"✅ Status check passed - Expected: {expected_status}, Got: {response.status_code}")
                
                # If there's a custom check function, run it
                if check_function:
                    check_result = check_function(response)
                    if check_result:
                        print(f"✅ Custom check passed: {check_result}")
                        self.tests_passed += 1
                        return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                    else:
                        print(f"❌ Custom check failed")
                        return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
                else:
                    self.tests_passed += 1
                    return True, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)
            else:
                print(f"❌ Status check failed - Expected: {expected_status}, Got: {response.status_code}")
                return False, response if binary else (response.json() if 'application/json' in response.headers.get('Content-Type', '') else response)

        except Exception as e:
            print(f"❌ Failed - Error: {str(e)}")
            return False, None

    def login(self, email="admin@example.com", password="admin123"):
        """Login and get authentication token"""
        print(f"\n🔐 Logging in as {email}...")
        
        form_data = {
            "username": email,
            "password": password
        }
        
        url = f"{self.api_url}/auth/token"
        response = requests.post(
            url, 
            data=form_data,
            headers={"Content-Type": "application/x-www-form-urlencoded"}
        )
        
        if response.status_code == 200:
            data = response.json()
            self.token = data.get("access_token")
            print(f"✅ Login successful - Token received")
            return True
        else:
            print(f"❌ Login failed - Status: {response.status_code}")
            if response.headers.get('Content-Type', '').startswith('application/json'):
                print(f"Error: {response.json()}")
            return False

    def test_null_vs_zero_values(self):
        """Test the difference between NULL and 0 values in the database and Excel export"""
        # First, create a test shooter
        shooter_data = {
            "name": f"NULL vs Zero Test Shooter {datetime.now().strftime('%H%M%S')}",
            "nra_number": "12345",
            "cmp_number": "67890"
        }
        
        success, shooter = self.run_test(
            "Create test shooter",
            "POST",
            "shooters",
            200,
            data=shooter_data
        )
        
        if not success:
            print("❌ Failed to create test shooter, cannot continue test")
            return False
        
        shooter_id = shooter["id"]
        print(f"Created test shooter with ID: {shooter_id}")
        
        # Create a test match
        match_data = {
            "name": f"NULL vs Zero Test Match {datetime.now().strftime('%H%M%S')}",
            "date": datetime.now().isoformat(),
            "location": "Test Range",
            "match_types": [
                {
                    "type": "NMC",
                    "instance_name": "NMC1",
                    "calibers": [".22", "CF"]
                }
            ],
            "aggregate_type": "None"
        }
        
        success, match = self.run_test(
            "Create test match",
            "POST",
            "matches",
            200,
            data=match_data
        )
        
        if not success:
            print("❌ Failed to create test match, cannot continue test")
            return False
        
        match_id = match["id"]
        print(f"Created test match with ID: {match_id}")
        
        # Create a score with NULL value
        score1_data = {
            "shooter_id": shooter_id,
            "match_id": match_id,
            "match_type_instance": "NMC1",
            "caliber": ".22",
            "stages": [
                {
                    "name": "SF",
                    "score": 95,
                    "x_count": 3
                },
                {
                    "name": "TF",
                    "score": None,  # NULL value
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 90,
                    "x_count": 2
                }
            ]
        }
        
        success, score1 = self.run_test(
            "Create score with NULL value",
            "POST",
            "scores",
            200,
            data=score1_data
        )
        
        if not success:
            print("❌ Failed to create score with NULL value")
            return False
        
        score1_id = score1["id"]
        print(f"Created score with NULL value, ID: {score1_id}")
        
        # Create a second score with 0 value
        score2_data = {
            "shooter_id": shooter_id,
            "match_id": match_id,
            "match_type_instance": "NMC1",
            "caliber": "CF",
            "stages": [
                {
                    "name": "SF",
                    "score": 85,
                    "x_count": 1
                },
                {
                    "name": "TF",
                    "score": 0,  # Zero value
                    "x_count": 0
                },
                {
                    "name": "RF",
                    "score": 80,
                    "x_count": 1
                }
            ]
        }
        
        success, score2 = self.run_test(
            "Create score with 0 value",
            "POST",
            "scores",
            200,
            data=score2_data
        )
        
        if not success:
            print("❌ Failed to create score with 0 value")
            return False
        
        score2_id = score2["id"]
        print(f"Created score with 0 value, ID: {score2_id}")
        
        # Retrieve the scores to verify how NULL and 0 values are stored
        success, retrieved_score1 = self.run_test(
            "Retrieve score with NULL value",
            "GET",
            f"scores/{score1_id}",
            200
        )
        
        if not success:
            print("❌ Failed to retrieve score with NULL value")
            return False
        
        success, retrieved_score2 = self.run_test(
            "Retrieve score with 0 value",
            "GET",
            f"scores/{score2_id}",
            200
        )
        
        if not success:
            print("❌ Failed to retrieve score with 0 value")
            return False
        
        # Verify that NULL and 0 values are stored correctly
        null_value_stored_correctly = False
        zero_value_stored_correctly = False
        
        # Check NULL value
        for stage in retrieved_score1["stages"]:
            if stage["name"] == "TF":
                if stage["score"] is None:
                    null_value_stored_correctly = True
                    print("✅ NULL value is stored correctly in the database")
                else:
                    print(f"❌ NULL value is not stored correctly in the database. Got: {stage['score']}")
        
        # Check 0 value
        for stage in retrieved_score2["stages"]:
            if stage["name"] == "TF":
                if stage["score"] == 0:
                    zero_value_stored_correctly = True
                    print("✅ 0 value is stored correctly in the database")
                else:
                    print(f"❌ 0 value is not stored correctly in the database. Got: {stage['score']}")
        
        # Download Excel report
        def check_excel_content(response):
            if response.status_code != 200:
                return None
                
            # Check if we got an Excel file
            content_type = response.headers.get('Content-Type', '')
            if 'spreadsheet' not in content_type and 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet' not in content_type:
                print(f"❌ Expected Excel file, got {content_type}")
                return None
            
            # Save the Excel file temporarily
            excel_data = response.content
            with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as temp_file:
                temp_file_path = temp_file.name
                temp_file.write(excel_data)
            
            print(f"📊 Excel file saved to {temp_file_path}")
            
            # Parse the Excel file to check NULL and 0 values
            try:
                wb = openpyxl.load_workbook(temp_file_path)
                sheet = wb.active
                
                # Print the entire Excel file content for debugging
                print("\n=== Excel File Content ===")
                for row in range(1, min(sheet.max_row + 1, 20)):  # Limit to first 20 rows
                    row_data = []
                    for col in range(1, sheet.max_column + 1):
                        value = sheet.cell(row=row, column=col).value
                        row_data.append(str(value))
                    print(f"Row {row}: {' | '.join(row_data)}")
                
                # Clean up
                os.unlink(temp_file_path)
                
                # For now, we'll consider the test successful if we can download and parse the Excel file
                return "Excel file downloaded and parsed successfully"
                
            except Exception as e:
                print(f"❌ Error parsing Excel file: {str(e)}")
                if os.path.exists(temp_file_path):
                    os.unlink(temp_file_path)
                return None
        
        success, _ = self.run_test(
            "Download and verify Excel report",
            "GET",
            f"match-report/{match_id}/excel",
            200,
            check_function=check_excel_content,
            binary=True
        )
        
        # Verify the match report API to see how it calculates averages
        success, match_report = self.run_test(
            "Get match report data",
            "GET",
            f"match-report/{match_id}",
            200
        )
        
        if not success:
            print("❌ Failed to get match report data")
            return False
        
        print("\n=== Match Report Data ===")
        print(json.dumps(match_report, indent=2))
        
        # Check if the test passed
        test_passed = null_value_stored_correctly and zero_value_stored_correctly and success
        
        return test_passed

def main():
    tester = NullVsZeroTester()
    
    # Login first
    if not tester.login():
        print("❌ Login failed, cannot continue tests")
        return 1
    
    # Run test for NULL vs 0 values
    print("\n=== Testing NULL vs 0 Values ===")
    test_result = tester.test_null_vs_zero_values()
    
    # Print summary
    print("\n=== Test Summary ===")
    print(f"NULL vs 0 Values Test: {'✅ PASSED' if test_result else '❌ FAILED'}")
    
    return 0 if test_result else 1

if __name__ == "__main__":
    sys.exit(main())

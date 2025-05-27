import requests
import json
import os
import io
import uuid
import sys
from datetime import datetime
import openpyxl
from openpyxl.styles import PatternFill, Alignment

# Get the backend URL from environment variable or use default
BACKEND_URL = os.environ.get("REACT_APP_BACKEND_URL", "http://localhost:8001")
API_URL = f"{BACKEND_URL}/api"

# Test credentials
TEST_EMAIL = "test@example.com"
TEST_PASSWORD = "testpassword"
TEST_USERNAME = "testuser"

# Global variables to store test data
auth_token = None
test_user_id = None
test_shooter_id = None
test_match_id = None
test_scores = []

def login():
    global auth_token, test_user_id
    
    # Check if user already exists
    try:
        response = requests.post(
            f"{API_URL}/auth/token",
            data={"username": TEST_EMAIL, "password": TEST_PASSWORD}
        )
        if response.status_code == 200:
            data = response.json()
            auth_token = data["access_token"]
            test_user_id = data["user_id"]
            print(f"Logged in existing user: {TEST_EMAIL}")
            return True
    except:
        pass
    
    # Register new user
    user_data = {
        "email": TEST_EMAIL,
        "username": TEST_USERNAME,
        "password": TEST_PASSWORD,
        "role": "admin"
    }
    
    response = requests.post(f"{API_URL}/auth/register", json=user_data)
    
    if response.status_code == 200:
        print(f"Registered new user: {TEST_EMAIL}")
        
        # Login to get token
        response = requests.post(
            f"{API_URL}/auth/token",
            data={"username": TEST_EMAIL, "password": TEST_PASSWORD}
        )
        
        if response.status_code == 200:
            data = response.json()
            auth_token = data["access_token"]
            test_user_id = data["user_id"]
            print(f"Logged in as: {TEST_EMAIL}")
            return True
        else:
            print(f"Login failed: {response.status_code} - {response.text}")
            return False
    else:
        print(f"Registration failed: {response.status_code} - {response.text}")
        return False

def create_test_shooter():
    global test_shooter_id, auth_token
    
    headers = {"Authorization": f"Bearer {auth_token}"}
    
    # Create a test shooter
    shooter_data = {
        "name": f"Header Alignment Test Shooter {uuid.uuid4().hex[:8]}",
        "nra_number": "12345678",
        "cmp_number": "87654321"
    }
    
    response = requests.post(f"{API_URL}/shooters", json=shooter_data, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        test_shooter_id = data["id"]
        print(f"Created test shooter: {shooter_data['name']} with ID: {test_shooter_id}")
        return True
    else:
        print(f"Failed to create shooter: {response.status_code} - {response.text}")
        return False

def create_test_match():
    global test_match_id, auth_token
    
    headers = {"Authorization": f"Bearer {auth_token}"}
    
    # Create a test match with multiple match types and calibers
    match_data = {
        "name": f"Header Alignment Test Match {uuid.uuid4().hex[:8]}",
        "date": datetime.now().isoformat(),
        "location": "Test Range",
        "match_types": [
            {
                "type": "NMC",
                "instance_name": "NMC1",
                "calibers": [".22", "CF", ".45"]
            },
            {
                "type": "600",
                "instance_name": "600_1",
                "calibers": [".22", "CF"]
            },
            {
                "type": "900",
                "instance_name": "900_1",
                "calibers": [".22"]
            },
            {
                "type": "Presidents",
                "instance_name": "Pres1",
                "calibers": ["Service Pistol", "Service Revolver"]
            }
        ],
        "aggregate_type": "None"
    }
    
    response = requests.post(f"{API_URL}/matches", json=match_data, headers=headers)
    
    if response.status_code == 200:
        data = response.json()
        test_match_id = data["id"]
        print(f"Created test match: {match_data['name']} with ID: {test_match_id}")
        return True
    else:
        print(f"Failed to create match: {response.status_code} - {response.text}")
        return False

def create_test_scores():
    global test_scores, auth_token, test_shooter_id, test_match_id
    
    headers = {"Authorization": f"Bearer {auth_token}"}
    
    # Create scores for different match types and calibers
    score_configs = [
        # NMC1 - .22
        {
            "caliber": ".22",
            "match_type_instance": "NMC1",
            "stages": [
                {"name": "SF", "score": 95, "x_count": 3},
                {"name": "TF", "score": 92, "x_count": 2},
                {"name": "RF", "score": 90, "x_count": 1}
            ]
        },
        # NMC1 - CF
        {
            "caliber": "CF",
            "match_type_instance": "NMC1",
            "stages": [
                {"name": "SF", "score": 85, "x_count": 1},
                {"name": "TF", "score": 82, "x_count": 0},
                {"name": "RF", "score": 80, "x_count": 0}
            ]
        },
        # NMC1 - .45
        {
            "caliber": ".45",
            "match_type_instance": "NMC1",
            "stages": [
                {"name": "SF", "score": 75, "x_count": 0},
                {"name": "TF", "score": 72, "x_count": 0},
                {"name": "RF", "score": 70, "x_count": 0}
            ]
        },
        # 600_1 - .22
        {
            "caliber": ".22",
            "match_type_instance": "600_1",
            "stages": [
                {"name": "SF1", "score": 98, "x_count": 4},
                {"name": "SF2", "score": 97, "x_count": 3},
                {"name": "TF1", "score": 96, "x_count": 2},
                {"name": "TF2", "score": 95, "x_count": 1},
                {"name": "RF1", "score": 94, "x_count": 0},
                {"name": "RF2", "score": 93, "x_count": 0}
            ]
        },
        # 600_1 - CF
        {
            "caliber": "CF",
            "match_type_instance": "600_1",
            "stages": [
                {"name": "SF1", "score": 88, "x_count": 1},
                {"name": "SF2", "score": 87, "x_count": 1},
                {"name": "TF1", "score": 86, "x_count": 0},
                {"name": "TF2", "score": 85, "x_count": 0},
                {"name": "RF1", "score": 84, "x_count": 0},
                {"name": "RF2", "score": 83, "x_count": 0}
            ]
        },
        # 900_1 - .22
        {
            "caliber": ".22",
            "match_type_instance": "900_1",
            "stages": [
                {"name": "SF1", "score": 98, "x_count": 4},
                {"name": "SF2", "score": 97, "x_count": 3},
                {"name": "TF1", "score": 96, "x_count": 2},
                {"name": "TF2", "score": 95, "x_count": 1},
                {"name": "RF1", "score": 94, "x_count": 0},
                {"name": "RF2", "score": 93, "x_count": 0},
                {"name": "SFNMC", "score": 95, "x_count": 2},
                {"name": "TFNMC", "score": 92, "x_count": 1},
                {"name": "RFNMC", "score": 90, "x_count": 0}
            ]
        },
        # Pres1 - Service Pistol
        {
            "caliber": "Service Pistol",
            "match_type_instance": "Pres1",
            "stages": [
                {"name": "SF1", "score": 95, "x_count": 2},
                {"name": "SF2", "score": 94, "x_count": 1},
                {"name": "TF", "score": 92, "x_count": 0},
                {"name": "RF", "score": 90, "x_count": 0}
            ]
        },
        # Pres1 - Service Revolver
        {
            "caliber": "Service Revolver",
            "match_type_instance": "Pres1",
            "stages": [
                {"name": "SF1", "score": 85, "x_count": 1},
                {"name": "SF2", "score": 84, "x_count": 0},
                {"name": "TF", "score": 82, "x_count": 0},
                {"name": "RF", "score": 80, "x_count": 0}
            ]
        }
    ]
    
    success = True
    
    for config in score_configs:
        score_data = {
            "shooter_id": test_shooter_id,
            "match_id": test_match_id,
            **config
        }
        
        response = requests.post(f"{API_URL}/scores", json=score_data, headers=headers)
        
        if response.status_code == 200:
            data = response.json()
            test_scores.append(data["id"])
            print(f"Created {config['match_type_instance']} - {config['caliber']} score with ID: {data['id']}")
        else:
            print(f"Failed to create {config['match_type_instance']} - {config['caliber']} score: {response.status_code} - {response.text}")
            success = False
    
    return success

def export_match_report_to_excel():
    global auth_token, test_match_id
    
    headers = {"Authorization": f"Bearer {auth_token}"}
    
    # Export match report to Excel
    response = requests.get(f"{API_URL}/match-report/{test_match_id}/excel", headers=headers)
    
    if response.status_code == 200:
        # Save the Excel file
        filename = "header_alignment_test.xlsx"
        with open(filename, "wb") as f:
            f.write(response.content)
        print(f"Exported match report to {filename}")
        return filename
    else:
        print(f"Failed to export match report: {response.status_code} - {response.text}")
        return None

def verify_header_alignment(filename):
    """
    Verify that match type/caliber headers are properly centered and aligned with the filled background
    """
    if not filename:
        print("No Excel file to verify")
        return False
    
    try:
        # Load the workbook
        wb = openpyxl.load_workbook(filename)
        
        # Skip the first sheet (Match Report) and check individual detail sheets
        sheet_names = wb.sheetnames
        
        if len(sheet_names) < 2:
            print("No individual detail sheets found in the Excel file")
            return False
        
        all_aligned = True
        issues = []
        
        # Check each individual detail sheet
        for sheet_name in sheet_names[1:]:
            ws = wb[sheet_name]
            print(f"\nChecking sheet: {sheet_name}")
            
            # Find all match type/caliber headers
            header_rows = []
            for row_idx, row in enumerate(ws.rows, 1):
                cell_value = ws.cell(row=row_idx, column=1).value
                if cell_value and (" - " in str(cell_value)) and any(x in str(cell_value) for x in ["NMC", "600", "900", "Pres"]):
                    header_rows.append(row_idx)
                    print(f"Found header at row {row_idx}: {cell_value}")
            
            # Check each header row
            for row_idx in header_rows:
                # Get the header cell
                header_cell = ws.cell(row=row_idx, column=1)
                header_text = header_cell.value
                
                # Check if the cell is part of a merged range
                merged = False
                merged_range = None
                for merged_cell_range in ws.merged_cells.ranges:
                    if header_cell.coordinate in merged_cell_range:
                        merged = True
                        merged_range = merged_cell_range
                        break
                
                if not merged:
                    issues.append(f"Sheet '{sheet_name}', Row {row_idx}: Header cell is not merged")
                    all_aligned = False
                    continue
                
                # Check if the header text is centered
                if header_cell.alignment.horizontal != 'center':
                    issues.append(f"Sheet '{sheet_name}', Row {row_idx}: Header text is not centered")
                    all_aligned = False
                    print(f"❌ Header text alignment: {header_cell.alignment.horizontal} (should be center)")
                else:
                    print(f"✅ Header text is centered")
                
                # Check if all cells in the merged range have the same background fill
                start_col, start_row, end_col, end_row = merged_range.bounds
                fill_color = None
                all_cells_filled = True
                
                print(f"Merged range: {merged_range} (columns {start_col}-{end_col})")
                
                for col in range(start_col, end_col + 1):
                    cell = ws.cell(row=row_idx, column=col)
                    
                    # Check if the cell has a fill
                    if not cell.fill or cell.fill.fill_type != "solid":
                        issues.append(f"Sheet '{sheet_name}', Row {row_idx}, Column {col}: Cell does not have a solid fill")
                        all_aligned = False
                        all_cells_filled = False
                        print(f"❌ Cell at column {col} does not have a solid fill")
                        continue
                    
                    # Check if all cells have the same fill color
                    if fill_color is None:
                        fill_color = cell.fill.start_color.rgb
                    elif cell.fill.start_color.rgb != fill_color:
                        issues.append(f"Sheet '{sheet_name}', Row {row_idx}, Column {col}: Cell has different fill color")
                        all_aligned = False
                        print(f"❌ Cell at column {col} has different fill color: {cell.fill.start_color.rgb} vs {fill_color}")
                
                if all_cells_filled:
                    print(f"✅ All cells in merged range have the same fill color: {fill_color}")
                
                # Check if the header text is properly aligned with the shaded area
                if merged and header_cell.alignment.horizontal == 'center':
                    print(f"✅ Header text '{header_text}' is properly centered across the merged range")
                else:
                    print(f"❌ Header text '{header_text}' is not properly centered across the merged range")
        
        if all_aligned:
            print("\n✅ All match type/caliber headers are properly centered and aligned with the filled background")
        else:
            print("\n❌ Some match type/caliber headers have alignment issues:")
            for issue in issues:
                print(f"  - {issue}")
        
        return all_aligned
    
    except Exception as e:
        print(f"Error verifying Excel header alignment: {str(e)}")
        return False

def run_tests():
    # Setup
    if not login():
        print("Login failed, cannot continue tests")
        return False
    
    if not create_test_shooter():
        print("Failed to create test shooter, cannot continue tests")
        return False
    
    if not create_test_match():
        print("Failed to create test match, cannot continue tests")
        return False
    
    if not create_test_scores():
        print("Failed to create test scores, cannot continue tests")
        return False
    
    # Export and verify
    excel_file = export_match_report_to_excel()
    if not excel_file:
        print("Failed to export match report, cannot continue tests")
        return False
    
    header_alignment_ok = verify_header_alignment(excel_file)
    
    # Print summary
    print("\n=== Test Summary ===")
    print(f"Excel Export: {'✅ Success' if excel_file else '❌ Failed'}")
    print(f"Header Alignment: {'✅ Correct' if header_alignment_ok else '❌ Incorrect'}")
    
    return header_alignment_ok

if __name__ == "__main__":
    success = run_tests()
    print(f"\nOverall test result: {'✅ PASSED' if success else '❌ FAILED'}")
    sys.exit(0 if success else 1)
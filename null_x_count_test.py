import requests
import json
import os
import uuid
import time
from datetime import datetime, timedelta
import io
import openpyxl

# Get the backend URL from the frontend .env file
def get_backend_url():
    with open('/app/frontend/.env', 'r') as f:
        for line in f:
            if line.startswith('REACT_APP_BACKEND_URL='):
                return line.strip().split('=')[1].strip('"\'')
    return None

BACKEND_URL = get_backend_url()
if not BACKEND_URL:
    raise ValueError("Could not find REACT_APP_BACKEND_URL in frontend/.env")

API_URL = f"{BACKEND_URL}/api"
AUTH_URL = f"{BACKEND_URL}/api/auth"

class TestNullXCountHandling:
    def __init__(self):
        self.token = None
        self.user_id = None
        self.shooter_id = None
        self.match_id = None
        self.score_ids = {}
        
    def setup(self):
        """Set up the test by logging in and creating test data"""
        print("\n=== Setting up test environment for NULL x_count testing ===")
        self.login()
        self.create_shooter()
        self.create_match()
        
    def login(self):
        """Log in as admin"""
        print("Logging in as admin...")
        response = requests.post(
            f"{AUTH_URL}/token",
            data={"username": "admin@example.com", "password": "admin123"}
        )
        
        if response.status_code == 200:
            data = response.json()
            self.token = data["access_token"]
            self.user_id = data["user_id"]
            print(f"Login successful. User ID: {self.user_id}")
        else:
            print(f"Login failed: {response.status_code} - {response.text}")
            raise Exception("Login failed")
            
    def get_headers(self):
        """Get headers with authentication token"""
        return {
            "Authorization": f"Bearer {self.token}",
            "Content-Type": "application/json"
        }
        
    def create_shooter(self):
        """Create a test shooter"""
        print("Creating test shooter...")
        shooter_data = {
            "name": f"X-Count Test Shooter {uuid.uuid4().hex[:8]}",
            "nra_number": "12345678",
            "cmp_number": "87654321"
        }
        
        response = requests.post(
            f"{API_URL}/shooters",
            headers=self.get_headers(),
            json=shooter_data
        )
        
        if response.status_code == 200:
            data = response.json()
            self.shooter_id = data["id"]
            print(f"Created shooter with ID: {self.shooter_id}")
        else:
            print(f"Failed to create shooter: {response.status_code} - {response.text}")
            raise Exception("Failed to create shooter")
            
    def create_match(self):
        """Create a test match with NMC and 600 match types"""
        print("Creating test match...")
        match_data = {
            "name": f"X-Count Test Match {uuid.uuid4().hex[:8]}",
            "date": (datetime.now() + timedelta(days=1)).isoformat(),
            "location": "Test Range",
            "match_types": [
                {
                    "type": "NMC",
                    "instance_name": "NMC1",
                    "calibers": [".22", "CF", ".45"]
                },
                {
                    "type": "600",
                    "instance_name": "600_1",
                    "calibers": [".22", "CF", ".45"]
                }
            ],
            "aggregate_type": "None"
        }
        
        response = requests.post(
            f"{API_URL}/matches",
            headers=self.get_headers(),
            json=match_data
        )
        
        if response.status_code == 200:
            data = response.json()
            self.match_id = data["id"]
            print(f"Created match with ID: {self.match_id}")
        else:
            print(f"Failed to create match: {response.status_code} - {response.text}")
            raise Exception("Failed to create match")
    
    def create_test_scores(self):
        """Create various test scores with different NULL x_count scenarios"""
        print("\n=== Creating test scores with NULL x_count values ===")
        
        # Test scenarios:
        # 1. Regular score with NULL x_count for all stages
        # 2. Regular score with mix of NULL and non-NULL x_count
        # 3. Zero score with NULL x_count
        # 4. NULL score with NULL x_count
        
        # Scenario 1: Regular score with NULL x_count for all stages (NMC)
        self.create_score("NMC1", ".22", [
            {"name": "SF", "score": 95, "x_count": None},
            {"name": "TF", "score": 92, "x_count": None},
            {"name": "RF", "score": 90, "x_count": None}
        ], "all_null_x")
        
        # Scenario 2: Regular score with mix of NULL and non-NULL x_count (NMC)
        self.create_score("NMC1", "CF", [
            {"name": "SF", "score": 88, "x_count": 2},
            {"name": "TF", "score": 85, "x_count": None},
            {"name": "RF", "score": 82, "x_count": 1}
        ], "mixed_null_x")
        
        # Scenario 3: Zero score with NULL x_count (NMC)
        self.create_score("NMC1", ".45", [
            {"name": "SF", "score": 0, "x_count": None},
            {"name": "TF", "score": 0, "x_count": None},
            {"name": "RF", "score": 0, "x_count": None}
        ], "zero_score_null_x")
        
        # Scenario 4: NULL score with NULL x_count (600)
        self.create_score("600_1", ".22", [
            {"name": "SF1", "score": None, "x_count": None},
            {"name": "SF2", "score": None, "x_count": None},
            {"name": "TF1", "score": None, "x_count": None},
            {"name": "TF2", "score": None, "x_count": None},
            {"name": "RF1", "score": None, "x_count": None},
            {"name": "RF2", "score": None, "x_count": None}
        ], "null_score_null_x")
        
        # Scenario 5: Mixed scores with NULL x_count (600)
        self.create_score("600_1", "CF", [
            {"name": "SF1", "score": 98, "x_count": None},
            {"name": "SF2", "score": 97, "x_count": 3},
            {"name": "TF1", "score": 96, "x_count": None},
            {"name": "TF2", "score": 95, "x_count": 2},
            {"name": "RF1", "score": 94, "x_count": None},
            {"name": "RF2", "score": 93, "x_count": 1}
        ], "mixed_scores_null_x")
        
    def create_score(self, match_type_instance, caliber, stages, scenario_name):
        """Create a score with the given parameters"""
        print(f"Creating {scenario_name} score for {match_type_instance} {caliber}...")
        
        score_data = {
            "shooter_id": self.shooter_id,
            "match_id": self.match_id,
            "caliber": caliber,
            "match_type_instance": match_type_instance,
            "stages": stages
        }
        
        response = requests.post(
            f"{API_URL}/scores",
            headers=self.get_headers(),
            json=score_data
        )
        
        if response.status_code == 200:
            data = response.json()
            self.score_ids[scenario_name] = data["id"]
            
            # Print the total score and x_count
            total_score = data.get("total_score")
            total_x_count = data.get("total_x_count")
            print(f"  Created score with ID: {data['id']}")
            print(f"  Total score: {total_score}, Total X count: {total_x_count}")
            
            # Verify that total_x_count is calculated correctly (sum of non-NULL x_count values)
            expected_x_count = sum(stage.get("x_count", 0) for stage in stages if stage.get("x_count") is not None)
            if total_x_count == expected_x_count:
                print(f"  ✅ Total X count calculated correctly: {total_x_count}")
            else:
                print(f"  ❌ Total X count incorrect. Expected: {expected_x_count}, Got: {total_x_count}")
        else:
            print(f"Failed to create score: {response.status_code} - {response.text}")
            raise Exception(f"Failed to create {scenario_name} score")
    
    def test_update_score_with_null_x_count(self):
        """Test updating a score with NULL x_count values"""
        print("\n=== Testing score update with NULL x_count values ===")
        
        # Update the "mixed_null_x" score to have more NULL x_count values
        score_id = self.score_ids.get("mixed_null_x")
        if not score_id:
            print("No score ID found for 'mixed_null_x'")
            return False
            
        print(f"Updating score {score_id} with NULL x_count values...")
        
        # Get the current score
        response = requests.get(
            f"{API_URL}/scores/{score_id}",
            headers=self.get_headers()
        )
        
        if response.status_code != 200:
            print(f"Failed to get score: {response.status_code} - {response.text}")
            return False
            
        score_data = response.json()
        
        # Update the score with NULL x_count values
        score_data["stages"][0]["x_count"] = None  # Set SF x_count to NULL
        
        # Remove id and created_at fields for the PUT request
        if "id" in score_data:
            del score_data["id"]
        if "created_at" in score_data:
            del score_data["created_at"]
        if "total_score" in score_data:
            del score_data["total_score"]
        if "total_x_count" in score_data:
            del score_data["total_x_count"]
            
        response = requests.put(
            f"{API_URL}/scores/{score_id}",
            headers=self.get_headers(),
            json=score_data
        )
        
        if response.status_code == 200:
            data = response.json()
            updated_total_score = data.get("total_score")
            updated_total_x_count = data.get("total_x_count")
            print(f"  Updated score successfully")
            print(f"  New total score: {updated_total_score}, New total X count: {updated_total_x_count}")
            
            # Verify that the total x_count excludes the NULL values
            expected_x_count = sum(stage.get("x_count", 0) for stage in score_data["stages"] if stage.get("x_count") is not None)
            
            if updated_total_x_count == expected_x_count:
                print(f"  ✅ Total X count calculated correctly (excluding NULL values): {updated_total_x_count}")
                return True
            else:
                print(f"  ❌ Total X count incorrect. Expected: {expected_x_count}, Got: {updated_total_x_count}")
                return False
        else:
            print(f"Failed to update score: {response.status_code} - {response.text}")
            return False
    
    def test_excel_export_x_count(self):
        """Test the Excel export functionality for NULL x_count handling"""
        print("\n=== Testing Excel export NULL x_count handling ===")
        
        # Get the Excel export
        response = requests.get(
            f"{API_URL}/match-report/{self.match_id}/excel",
            headers=self.get_headers()
        )
        
        if response.status_code != 200:
            print(f"Failed to get Excel export: {response.status_code} - {response.text}")
            return False
            
        # Save the Excel file temporarily
        excel_data = response.content
        excel_file = io.BytesIO(excel_data)
        
        # Load the workbook
        wb = openpyxl.load_workbook(excel_file)
        
        # Check the Match Report sheet
        print("Checking Match Report sheet...")
        ws = wb["Match Report"]
        
        # Find the row with our shooter
        shooter_row = None
        for row in range(9, ws.max_row + 1):  # Start from row 9 (after headers)
            cell_value = ws.cell(row=row, column=1).value
            if cell_value and "X-Count Test Shooter" in cell_value:
                shooter_row = row
                break
                
        if not shooter_row:
            print("  ❌ Could not find shooter in Match Report sheet")
            return False
            
        # Check NULL x_count handling in the summary sheet
        print("  Checking NULL x_count handling in summary sheet...")
        
        # Find columns for each match type and caliber
        column_headers = {}
        for col in range(3, ws.max_column + 1):  # Start from column 3 (after Shooter and Average)
            header = ws.cell(row=8, column=col).value
            if header:
                column_headers[header] = col
                
        # Check NULL x_count display for .22 NMC1 (all NULL x_count)
        nmc_22_header = "NMC1 (.22)"
        if nmc_22_header in column_headers:
            cell = ws.cell(row=shooter_row, column=column_headers[nmc_22_header]).value
            if cell and "X" not in str(cell):
                print(f"  ✅ NULL x_count displayed correctly without X count in summary for NMC1 .22: {cell}")
            else:
                print(f"  ❌ NULL x_count not displayed correctly in summary. Got: {cell}")
                return False
        else:
            print(f"  ❌ Could not find column for {nmc_22_header}")
            return False
            
        # Check NULL x_count display for CF NMC1 (mixed NULL and non-NULL x_count)
        nmc_cf_header = "NMC1 (CF)"
        if nmc_cf_header in column_headers:
            cell = ws.cell(row=shooter_row, column=column_headers[nmc_cf_header]).value
            if cell and "X" in str(cell):
                print(f"  ✅ Mixed NULL x_count displayed correctly with X count in summary for NMC1 CF: {cell}")
            else:
                print(f"  ❌ Mixed NULL x_count not displayed correctly in summary. Got: {cell}")
                return False
        else:
            print(f"  ❌ Could not find column for {nmc_cf_header}")
            return False
            
        # Check individual shooter detail sheet
        print("Checking shooter detail sheet...")
        
        # Find the shooter's detail sheet
        shooter_sheet = None
        for sheet_name in wb.sheetnames:
            if "X-Count Test Shooter" in sheet_name:
                shooter_sheet = wb[sheet_name]
                break
                
        if not shooter_sheet:
            print("  ❌ Could not find shooter detail sheet")
            return False
            
        # Check NULL x_count display in detail sheet
        null_x_count_found = False
        mixed_x_count_found = False
        
        for row in range(1, shooter_sheet.max_row + 1):
            cell_value = shooter_sheet.cell(row=row, column=1).value
            
            # Check for NULL x_count section (NMC1 .22)
            if cell_value == "NMC1 - .22":
                # Check the stages below this header
                for stage_row in range(row + 2, row + 10):  # Check a few rows
                    stage_name = shooter_sheet.cell(row=stage_row, column=1).value
                    if stage_name in ["SF", "TF", "RF"]:
                        score_value = shooter_sheet.cell(row=stage_row, column=2).value
                        x_count_value = shooter_sheet.cell(row=stage_row, column=3).value
                        
                        if score_value is not None and x_count_value == "-":
                            null_x_count_found = True
                            print(f"  ✅ NULL x_count displayed as '-' in detail sheet for stage {stage_name}")
                            
            # Check for mixed NULL x_count section (NMC1 CF)
            if cell_value == "NMC1 - CF":
                # Check the stages below this header
                for stage_row in range(row + 2, row + 10):  # Check a few rows
                    stage_name = shooter_sheet.cell(row=stage_row, column=1).value
                    if stage_name == "TF":
                        score_value = shooter_sheet.cell(row=stage_row, column=2).value
                        x_count_value = shooter_sheet.cell(row=stage_row, column=3).value
                        
                        if score_value is not None and x_count_value == "-":
                            mixed_x_count_found = True
                            print(f"  ✅ NULL x_count displayed as '-' in detail sheet for stage {stage_name}")
        
        if not null_x_count_found:
            print("  ❌ Could not find NULL x_count display in detail sheet")
            return False
            
        if not mixed_x_count_found:
            print("  ❌ Could not find mixed NULL x_count display in detail sheet")
            return False
            
        print("  ✅ Excel export correctly handles NULL x_count values")
        return True
        
    def run_tests(self):
        """Run all tests"""
        try:
            self.setup()
            self.create_test_scores()
            
            # Test updating scores with NULL x_count values
            update_result = self.test_update_score_with_null_x_count()
            
            # Test Excel export
            excel_result = self.test_excel_export_x_count()
            
            # Print overall results
            print("\n=== Test Results for NULL x_count handling ===")
            print(f"Score update with NULL x_count values: {'✅ PASS' if update_result else '❌ FAIL'}")
            print(f"Excel export NULL x_count handling: {'✅ PASS' if excel_result else '❌ FAIL'}")
            
            overall_result = update_result and excel_result
            print(f"\nOverall result: {'✅ PASS' if overall_result else '❌ FAIL'}")
            
            return overall_result
            
        except Exception as e:
            print(f"Test failed with exception: {str(e)}")
            return False


if __name__ == "__main__":
    test = TestNullXCountHandling()
    test.run_tests()